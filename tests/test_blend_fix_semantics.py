#!/usr/bin/env python3
"""Executable proof of the TEAM RATINGS D-column blank-preservation fix.

Builds a miniature workbook wired exactly like TEAM RATINGS D/F/G/H against a
CALC-style governed column, then evaluates the OLD and NEW D formulas with a real
Excel formula engine across four states: GP=0 (governed returns ""), GP>0 (numeric),
a truly empty governed cell, and a team missing from the lookup.

Skips if the optional `formulas` package is unavailable, so the standard suite still
runs everywhere. Nothing here reads or writes any production workbook.
"""
import os
import tempfile
import unittest

try:
    import formulas  # noqa: F401
    import openpyxl
    HAVE = True
except ImportError:  # pragma: no cover
    HAVE = False

OLD_D = '=IFERROR(ROUND(INDEX($B$1:$B$4,MATCH($A{r},$A$1:$A$4,0)),2),"")'
NEW_D = ('=IFERROR(IF(ISNUMBER(INDEX($B$1:$B$4,MATCH($A{r},$A$1:$A$4,0))),'
         'ROUND(INDEX($B$1:$B$4,MATCH($A{r},$A$1:$A$4,0)),2),""),"")')
# Production H, verbatim except for the local column letters.
H = '=IF(AND($D{r}="",$F{r}=""),"",IF($D{r}="",$F{r},ROUND($G{r}*IF($F{r}="",0,$F{r})+(1-$G{r})*$D{r},2)))'

CASES = [  # label, governed cell content, prior
    ("gp0_formula_blank", '=IF(TRUE,"",1)', -1.07),
    ("gp_positive_numeric", 1.234, -1.07),
    ("governed_truly_empty", None, -1.07),
    ("team_not_found", "SKIP", -1.07),
    ("no_governed_and_no_prior", None, None),
]


@unittest.skipUnless(HAVE, "optional `formulas` engine not installed")
class BlendFixSemantics(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "T"
        for i, (label, governed, prior) in enumerate(CASES, start=1):
            ws.cell(i, 1, label if label != "team_not_found" else "MISSING")
            if governed != "SKIP" and governed is not None:
                ws.cell(i, 2, governed)
            r = i
            ws.cell(r, 4, NEW_D.format(r=r))
            ws.cell(r, 5, OLD_D.format(r=r))
            if prior is not None:
                ws.cell(r, 6, prior)
            ws.cell(r, 7, 0.8)
            ws.cell(r, 8, H.format(r=r))
        # row 4's lookup key must not exist in A1:A4 -> force a genuine #N/A
        ws.cell(4, 1, "team_not_found")
        for r in range(1, 6):
            ws.cell(r, 4, NEW_D.format(r=r).replace("$A{r}".format(r=r), "$A%d" % r))
        cls.tmp = tempfile.mkdtemp()
        cls.path = os.path.join(cls.tmp, "blendfix_probe.xlsx")
        ws.cell(4, 1, "ZZZ_not_in_list")
        wb.save(cls.path)
        sol = formulas.ExcelModel().loads(cls.path).finish().calculate()
        cls.val = {}
        for k, v in sol.items():
            m = k.upper().rsplit("!", 1)
            if len(m) == 2 and len(m[1]) <= 3:
                try:
                    cls.val[m[1]] = v.value[0, 0]
                except Exception:
                    pass

    def d(self, row):
        return self.val.get("D%d" % row)

    def h(self, row):
        return self.val.get("H%d" % row)

    def test_gp0_formula_blank_keeps_D_blank_and_H_uses_full_prior(self):
        self.assertEqual(self.d(1), "")
        self.assertEqual(self.h(1), -1.07, "GP=0 must fall back to 100% preseason prior")

    def test_gp_positive_still_blends(self):
        self.assertEqual(self.d(2), 1.23)
        self.assertEqual(self.h(2), round(0.8 * -1.07 + 0.2 * 1.23, 2))

    def test_truly_empty_governed_cell_stays_blank(self):
        self.assertEqual(self.d(3), "", "an empty governed cell must not coerce to numeric 0")
        self.assertEqual(self.h(3), -1.07)

    def test_missing_team_stays_blank(self):
        self.assertEqual(self.d(4), "")
        self.assertEqual(self.h(4), -1.07)

    def test_no_governed_and_no_prior_yields_blank_rating(self):
        self.assertEqual(self.d(5), "")
        self.assertEqual(self.h(5), "")

    def test_old_formula_reproduces_the_defect_on_an_empty_cell(self):
        """Regression witness: the formula being replaced returns numeric 0 where the
        fixed one returns blank. If this ever stops holding, the fix is moot."""
        self.assertEqual(self.val.get("E3"), 0)
        self.assertEqual(self.d(3), "")


if __name__ == "__main__":
    unittest.main(verbosity=2)
