#!/usr/bin/env python3
"""Repo-local regression suite (stdlib only). Run: python3 tests/run_tests.py"""
import unittest, subprocess, csv, hashlib, os, sys, glob
ROOT=os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
os.chdir(ROOT)
AUTH="TTW_NFL_Power_Ratings_2026_v1.4_AUTHORITATIVE.xlsx"
# Pinned SHA of the COMMITTED LOCAL artifact. This is NOT a live-drift test: Google
# repackages every export, so a fresh export of the same unchanged Sheet has a different
# ZIP SHA. Live drift is measured semantically -- see scripts/semantic_fingerprint.py.
AUTH_SHA="39d42aa4863422e0b6df30553e1eb635cd451c9b9ecf63d2272acd78ed2a7fa9"
AUTH_FINGERPRINT="b6169953b71eed19579d974a399b0f431eb326d7a7923bd6c03847b591ee183b"

def sh(cmd): return subprocess.run(cmd,shell=True,capture_output=True,text=True)

class Gates(unittest.TestCase):
    def test_gate_suite_passes(self):
        r=sh("python3 scripts/run_gates.py"); self.assertEqual(r.returncode,0,r.stdout+r.stderr)
    def test_validator_passes(self):
        r=sh("python3 scripts/validate_preseason_monitor.py"); self.assertEqual(r.returncode,0,r.stdout+r.stderr)
    def test_linkcheck_passes(self):
        r=sh("python3 scripts/linkcheck_preseason.py"); self.assertEqual(r.returncode,0,r.stdout+r.stderr)

class Authoritative(unittest.TestCase):
    def test_authoritative_workbook_unmodified(self):
        self.assertEqual(hashlib.sha256(open(AUTH,'rb').read()).hexdigest(),AUTH_SHA,
                         "committed AUTHORITATIVE artifact must not be modified")
    def test_authoritative_semantic_fingerprint_pinned(self):
        """The drift test that survives Google repackaging."""
        sys.path.insert(0,os.path.join(ROOT,"scripts"))
        import semantic_fingerprint as SF
        fp=SF.fingerprint(AUTH)
        self.assertEqual(fp["digests"]["combined"],AUTH_FINGERPRINT)
        self.assertEqual(len(fp["formulas"]),57399)
        self.assertEqual(len(fp["sheets"]),22)
    def test_single_active_authoritative_workbook(self):
        act=[p for p in glob.glob("*.xlsx") if "AUTHORITATIVE" in p]
        self.assertEqual(act,[AUTH],"exactly one authoritative workbook may sit in the root")
    def test_no_new_spreadsheet_added(self):
        allowed={"TTW_NFL_Power_Ratings_2026_v1.4_AUTHORITATIVE.xlsx","TTW_NFL_v1_1_1 Version 2.xlsx",
                 "TTW_NFL_Power_Ratings_2026_v1.2_QB_WORKING.xlsx","TTW_NFL_Power_Ratings_2026_v1.2.1_QB_CANDIDATE.xlsx",
                 "TTW_NFL_Power_Ratings_2026_v1.2.2_QB_SWEEP_CANDIDATE.xlsx","TTW_NFL_Power_Ratings_2026_v1.3_MARKET_CANDIDATE.xlsx",
                 # Phase 3 (2026-09-01): read-only live-Sheet export used as the candidate base,
                 # and the Source-B + blend-fix candidate built from it. Both are candidate/
                 # provenance artifacts; neither is authoritative and neither is promoted.
                 "TTW_NFL_2026_LIVE_EXPORT_20260901_BASE.xlsx",
                 "TTW_NFL_Power_Ratings_2026_v1.4_SRCB_BLENDFIX_CANDIDATE.xlsx",
                 # Phase 5C (2026-09-01): pre-promotion production rollback checkpoint.
                 # Read-only capture; production was never written.
                 "TTW_NFL_2026_PROD_ROLLBACK_CHECKPOINT_20260901T1432Z.xlsx",
                 # Phase 5C post-promotion: read-only export of live production at v1.4.
                 "TTW_NFL_2026_PROD_POSTPROMOTION_20260901T1650Z.xlsx"}
        found={os.path.basename(p) for p in glob.glob("**/*.xls*",recursive=True)}
        self.assertEqual(found-allowed,set(),"no spreadsheet file may be added to the repository")

class Monitor(unittest.TestCase):
    def setUp(self): self.rows=list(csv.DictReader(open("preseason/PRESEASON_MONITOR.csv")))
    def test_32_team_game_rows(self):
        self.assertEqual(len(self.rows),32); self.assertEqual(len({r["Team"] for r in self.rows}),32)
    def test_current_tracking_state(self):
        expected={"MIN":("UPDATE","Y"),"CLE":("UPDATE","Y"),"NO":("UPDATE","Y"),"WAS":("MONITOR","N"),"CHI":("MONITOR","N")}
        for r in self.rows:
            got=(r["Decision"],r["Workbook Updated?"])
            self.assertEqual(got, expected.get(r["Team"],("PENDING","N")), r["Team"])
    def test_decision_updated_lifecycle_consistency(self):
        for r in self.rows:
            d,u=r["Decision"],r["Workbook Updated?"]
            self.assertIn(d,{"PENDING","MONITOR","UPDATE","IGNORE"},r["Team"])
            self.assertIn(u,{"Y","N"},r["Team"])
            if d in {"PENDING","MONITOR","IGNORE"}:
                self.assertEqual(u,"N",f"{r['Team']}: {d} must have N")
            self.assertFalse(d=="PENDING" and u=="Y",f"{r['Team']}: PENDING/Y forbidden")
    def test_unplayed_games_are_tbd(self):
        for r in self.rows:
            if r["Game Status"]=="NOT PLAYED": self.assertEqual(r["Starter Use"],"TBD",r["Team"])
    def test_complete_tbd_rows_have_blockers(self):
        for r in self.rows:
            if r["Game Status"]=="COMPLETE" and r["Starter Use"]=="TBD":
                self.assertTrue(r["Blocker"].strip(),f"{r['Team']}: COMPLETE+TBD needs a named blocker")
    def test_all_16_games_present(self):
        games={tuple(sorted([r["Team"],r["Opponent"]])) for r in self.rows}
        self.assertEqual(len(games),16,"expected 16 distinct games")
    def test_no_point_values_in_proposals(self):
        import re
        for r in self.rows:
            self.assertIsNone(re.search(r"[+-]?\d+(\.\d+)?\s*(pt|point)",r["Proposed Change"],re.I),r["Team"])

class LinkcheckMultiURL(unittest.TestCase):
    """Source URL may hold several URLs separated by ' ; '. Every one must be validated."""
    def setUp(self):
        sys.path.insert(0,os.path.join(ROOT,"scripts"))
        import linkcheck_preseason as L; self.L=L
    def test_multi_url_field_splits_into_both_urls(self):
        f="https://www.vikings.com/a ; https://www.neworleanssaints.com/b"
        self.assertEqual(self.L.split_source_urls(f),
                         ["https://www.vikings.com/a","https://www.neworleanssaints.com/b"])
    def test_both_urls_in_valid_multi_field_are_checked(self):
        for u in self.L.split_source_urls("https://www.vikings.com/a ; https://www.nfl.com/b"):
            self.assertTrue(self.L.check_url(u)[0], u)
    def test_non_allowlisted_second_url_is_detected(self):
        parts=self.L.split_source_urls("https://www.nfl.com/a ; https://evil.example.com/b")
        self.assertTrue(self.L.check_url(parts[0])[0])
        ok,host,reason=self.L.check_url(parts[1])
        self.assertFalse(ok); self.assertEqual(host,"evil.example.com"); self.assertIn("allowlist",reason)
    def test_invalid_second_url_is_detected(self):
        parts=self.L.split_source_urls("https://www.nfl.com/a ; http://www.nfl.com/b")
        ok,_h,reason=self.L.check_url(parts[1])
        self.assertFalse(ok); self.assertIn("non-HTTPS", reason)
    def test_single_url_behaviour_preserved(self):
        f="https://www.nfl.com/news/x"
        self.assertEqual(self.L.split_source_urls(f),[f])
        self.assertTrue(self.L.check_url(f)[0])
    def test_joined_field_is_never_treated_as_one_url(self):
        joined="https://www.nfl.com/a ; https://www.vikings.com/b"
        self.assertFalse(self.L.check_url(joined)[0])          # combined string is not a valid URL
        self.assertEqual(len(self.L.split_source_urls(joined)),2)

class DQGuardPatch(unittest.TestCase):
    """Proves the PROPOSED CALC guard semantics. No workbook is modified."""
    @staticmethod
    def guarded_mean(vals):
        nums=[v for v in vals if isinstance(v,(int,float))]
        return "" if len(nums)==0 else round(sum(nums)/len(nums),6)
    def test_empty_range_yields_blank_not_div0(self):
        self.assertEqual(self.guarded_mean(["","","" ]),"")
    def test_populated_range_matches_average(self):
        self.assertEqual(self.guarded_mean([1.0,2.0,3.0]),2.0)
    def test_mixed_ignores_text_blanks(self):
        self.assertEqual(self.guarded_mean(["",2.0,"",4.0]),3.0)
    def test_guard_preserves_nonzero_signal(self):
        self.assertNotEqual(self.guarded_mean([0.5,0.5]),"")

class ShadowAudit20260901(unittest.TestCase):
    """2026-09-01 preseason-prior shadow audit: artifacts regenerate byte-identically
    from the committed provenance JSON, and the headline invariants hold.
    Shadow-only: nothing here touches the authoritative workbook or the live Sheet."""
    PROV=os.path.join(ROOT,"audit","TTW_NFL_2026_Preseason_Prior_Provenance_20260901.json")
    SHADOW=os.path.join(ROOT,"audit","TTW_NFL_2026_Preseason_Prior_Shadow_20260901.csv")
    WK1=os.path.join(ROOT,"audit","TTW_NFL_2026_Week1_Before_After_Shadow_20260901.csv")
    def test_provenance_parses_with_32_teams_everywhere(self):
        import json
        p=json.load(open(self.PROV))
        self.assertEqual(len(p["source_A"]["values_raw"]),32)
        self.assertEqual(len(p["source_B_vsin"]["values"]),32)
        self.assertEqual(len(p["source_B_fpi"]["values"]),32)
        for b in ("betmgm","fanatics","draftkings_jul23"):
            self.assertEqual(len(p["source_C_win_totals"]["books"][b]["totals"]),32)
        self.assertEqual(len(p["market_lines_week1"]["games"]),16)
    def test_vsin_league_mean_exactly_24(self):
        import json
        v=json.load(open(self.PROV))["source_B_vsin"]["values"]
        self.assertAlmostEqual(sum(v.values())/32,24.0,places=9)
        self.assertEqual(v["DAL"],24.0); self.assertEqual(v["NYG"],22.0)
    def test_generator_reproduces_both_csvs_byte_identically(self):
        import subprocess,tempfile,shutil
        with tempfile.TemporaryDirectory() as td:
            for f in (self.SHADOW,self.WK1):
                shutil.copy(f,os.path.join(td,os.path.basename(f)))
            subprocess.run([sys.executable,os.path.join(ROOT,"scripts","gen_preseason_prior_shadow.py")],
                           check=True,capture_output=True)
            for f in (self.SHADOW,self.WK1):
                with open(f,"rb") as a, open(os.path.join(td,os.path.basename(f)),"rb") as b:
                    self.assertEqual(a.read(),b.read(),f"{os.path.basename(f)} not byte-identical")
    def test_shadow_csv_pins_current_production_chain(self):
        import csv
        rows={r["Team"]:r for r in csv.DictReader(open(self.SHADOW))}
        self.assertEqual(len(rows),32)
        self.assertEqual(rows["DAL"]["SrcA regressed"],"-2.82")
        self.assertEqual(rows["NYG"]["SrcA regressed"],"-1.99")
        self.assertEqual(rows["DAL"]["Wk1 eff current (W1-A)"],"-2.26")
        self.assertEqual(rows["NYG"]["Wk1 eff current (W1-A)"],"-1.59")
        self.assertEqual(rows["DAL"]["Rank current"],"27")
        self.assertEqual(rows["NYG"]["Rank current"],"24")
    def test_wk1_csv_pins_dal_nyg_edge_chain(self):
        import csv
        g={r["GameID"]:r for r in csv.DictReader(open(self.WK1))}
        r=g["2026_01_DAL_NYG"]
        self.assertEqual(r["Market home spread"],"2.5")
        self.assertEqual(r["current_W1A FinalMargin"],"2.27")
        self.assertEqual(r["current_W1A edge"],"4.77")
        self.assertEqual(r["current_W1A side"],"NYG")
    def test_authoritative_workbook_untouched_by_shadow_audit(self):
        import hashlib
        h=hashlib.sha256(open(os.path.join(ROOT,
            AUTH),"rb").read()).hexdigest()
        self.assertEqual(h,AUTH_SHA)

class CandidateSrcBBlendFix(unittest.TestCase):
    """Phase 3 candidate (2026-09-01): Source B populated + TEAM RATINGS blank-preservation
    fix, built from the committed read-only live export. Candidate only — never promoted."""
    BASE="TTW_NFL_2026_LIVE_EXPORT_20260901_BASE.xlsx"
    BASE_SHA="39c7c567364234068245e68d0af943ca71e4a128b652f57cec258d40ac1e3f35"
    CAND="TTW_NFL_Power_Ratings_2026_v1.4_SRCB_BLENDFIX_CANDIDATE.xlsx"
    CAND_SHA="a71e8ba3356fe456d678eb2db75ec67ddfdbe287e743cc9901600cf57c97e22e"
    SLATE="audit/TTW_NFL_2026_Candidate_Week1_Slate_20260901.csv"
    MANIFEST="audit/TTW_NFL_2026_Candidate_Changed_Cells_20260901.csv"
    def test_base_and_candidate_pinned_by_sha(self):
        for p,s in ((self.BASE,self.BASE_SHA),(self.CAND,self.CAND_SHA)):
            self.assertEqual(hashlib.sha256(open(p,'rb').read()).hexdigest(),s,p)
    def test_full_verification_suite_passes(self):
        r=sh(f"python3 scripts/verify_srcb_blendfix_candidate.py {self.BASE}")
        self.assertEqual(r.returncode,0,r.stdout+r.stderr)
        self.assertIn("0 failed",r.stdout)
    def test_candidate_rebuilds_byte_identically(self):
        import shutil,tempfile
        with tempfile.TemporaryDirectory() as td:
            keep=os.path.join(td,"cand.xlsx"); shutil.copy(self.CAND,keep)
            r=sh(f"python3 scripts/build_srcb_blendfix_candidate.py {self.BASE}")
            self.assertEqual(r.returncode,0,r.stdout+r.stderr)
            self.assertEqual(open(self.CAND,'rb').read(),open(keep,'rb').read(),
                             "candidate must rebuild byte-identically from the committed base")
    def test_slate_pins_dal_nyg(self):
        rows={r["GameID"]:r for r in csv.DictReader(open(self.SLATE))}
        self.assertEqual(len(rows),16)
        g=rows["2026_01_DAL_NYG"]
        self.assertEqual(g["Away eff"],"-1.07"); self.assertEqual(g["Home eff"],"-1.9")
        self.assertEqual(g["FinalMargin"],"0.77")
        self.assertEqual(g["Model spread (fair line)"],"NYG -0.8")
        self.assertEqual(g["SpreadEdge"],"3.27")
        self.assertTrue(g["Supported side"].startswith("NYG"))
    def test_manifest_covers_every_changed_cell(self):
        rows=list(csv.DictReader(open(self.MANIFEST)))
        self.assertEqual(len(rows),32*11,"32 teams x 11 cells")
        self.assertEqual(len({r["Team"] for r in rows}),32)
        self.assertEqual(sum(1 for r in rows if "FORMULA CHANGED" in r["Kind"]),32)
    def test_v14_banner_and_changelog_entry(self):
        import openpyxl
        wb=openpyxl.load_workbook(self.CAND,data_only=True)
        self.assertEqual(wb["START HERE"]["A1"].value,
                         "TO THE WINDOW \u2014 NFL POWER RATINGS 2026 (v1.4)")
        ch=wb["CHANGELOG"]
        self.assertEqual(ch.cell(7,1).value,"1.4")
        self.assertEqual(ch.cell(7,2).value,"2026-09-01")
        for topic in ("Source B","all 32 teams","ISNUMBER","TEAM RATINGS!D5:D36",
                      "Source C","VALIDATE-ONLY","weights","thresholds","QB values",
                      "market lines","overrides","adjustments"):
            self.assertIn(topic,ch.cell(7,3).value,topic)
        self.assertIsNone(ch.cell(8,1).value,"exactly one new CHANGELOG row")
        wb.close()
    def test_authoritative_and_live_base_untouched_by_phase3(self):
        self.assertEqual(hashlib.sha256(open(AUTH,'rb').read()).hexdigest(),AUTH_SHA)

class ProductionPreflight5C(unittest.TestCase):
    """Phase 5C: production was read-only preflighted and found at zero drift from the
    promotion-manifest baseline. No write was performed (the connector has no cell-level
    write); production stays untouched."""
    CKPT="TTW_NFL_2026_PROD_ROLLBACK_CHECKPOINT_20260901T1432Z.xlsx"
    CKPT_SHA="e3349d8ee42fedae6cc411e9ca92f68bc6f7cd9c2ae2f378393bdc2d28199ce9"
    def test_rollback_checkpoint_pinned(self):
        self.assertEqual(hashlib.sha256(open(self.CKPT,'rb').read()).hexdigest(),self.CKPT_SHA)
    def test_preflight_reports_zero_drift(self):
        r=sh(f"python3 scripts/preflight_production_5c.py {self.CKPT}")
        self.assertEqual(r.returncode,0,r.stdout+r.stderr)
        self.assertIn("0 failed",r.stdout)
    def test_checkpoint_is_pre_promotion_state(self):
        import openpyxl
        wb=openpyxl.load_workbook(self.CKPT,data_only=True)
        tr,ps=wb["TEAM RATINGS"],wb["PRESEASON"]
        self.assertTrue(all(tr.cell(r,4).value==0 for r in range(5,37)),
                        "D column must still show the pre-fix coercion to 0")
        self.assertTrue(all(ps.cell(r,9).value is None for r in range(5,37)),
                        "Source B must still be empty in production")
        self.assertEqual(wb["START HERE"]["A1"].value,
                         "TO THE WINDOW \u2014 NFL POWER RATINGS 2026 (v1.1)")
        wb.close()
    def test_manual_apply_blocks_reproduce_the_candidate(self):
        import openpyxl,csv as _csv
        C=openpyxl.load_workbook("TTW_NFL_Power_Ratings_2026_v1.4_SRCB_BLENDFIX_CANDIDATE.xlsx",
                                 data_only=True)
        ps,ch=C["PRESEASON"],C["CHANGELOG"]
        D="promotion/manual_apply_v14"
        b1=[l.strip() for l in open(f"{D}/block1_PRESEASON_I5_I36.tsv")]
        self.assertEqual(len(b1),32)
        for i in range(32):
            self.assertAlmostEqual(float(b1[i]),ps.cell(5+i,9).value,places=9)
        b2=list(_csv.reader(open(f"{D}/block2_PRESEASON_K5_L36.tsv"),delimiter="\t"))
        self.assertEqual(len(b2),32)
        for i in range(32):
            self.assertEqual(b2[i][0],ps.cell(5+i,11).value)
            self.assertEqual(b2[i][1],ps.cell(5+i,12).value)
        b5=list(_csv.reader(open(f"{D}/block5_CHANGELOG_A7_D7.tsv"),delimiter="\t"))[0]
        for c in range(1,5):
            self.assertEqual(b5[c-1],ch.cell(7,c).value)
        self.assertIn("ISNUMBER",open(f"{D}/block3_TEAMRATINGS_D5_formula.txt").read())
        self.assertIn("(v1.4)",open(f"{D}/block4_STARTHERE_A1_banner.txt").read())
        C.close()

class SheetsProbeEvidence(unittest.TestCase):
    """Phase 4: values executed by Google Sheets in a disposable probe Sheet
    (1inj6XOyeCZflxkguPaMEyfhlNAjul5wBXEkZvl9r-pY, 2026-09-01). Pins the native
    evidence so a later edit to the shadow/candidate cannot silently diverge from it."""
    PROBE="audit/TTW_NFL_2026_Sheets_Probe_Results_20260901.csv"
    def setUp(self):
        self.rows=list(csv.reader(open(self.PROBE)))
    def test_block1_defect_and_fix_as_google_evaluates_them(self):
        b={r[0]:r for r in self.rows[2:6]}
        # old formula coerces a blank governed value to 0 -> H collapses to 80% of prior
        self.assertEqual(b["GP0"][4],"0");    self.assertEqual(b["GP0"][6],"-0.86")
        self.assertEqual(b["EMPTY"][4],"0");  self.assertEqual(b["EMPTY"][6],"-0.86")
        # fixed formula stays blank -> H falls back to 100% preseason prior
        for k in ("GP0","EMPTY","MISS"):
            self.assertEqual(b[k][5],"",k); self.assertEqual(b[k][7],"-1.07",k)
        # GP>0 still blends normally through both formulas
        self.assertEqual(b["GPPOS"][5],"1.23"); self.assertEqual(b["GPPOS"][7],"-0.61")
        self.assertEqual([b[k][9] for k in ("GP0","GPPOS","EMPTY","MISS")],
                         ["FALSE","TRUE","FALSE","FALSE"])
    def test_block2_matches_candidate_and_shadow_for_all_32(self):
        shadow={r["Team"]:r for r in csv.DictReader(
            open("audit/TTW_NFL_2026_Preseason_Prior_Shadow_20260901.csv"))}
        rows=self.rows[9:41]
        self.assertEqual(len(rows),32)
        for r in rows:
            t=r[0]
            self.assertEqual(float(r[3]),float(shadow[t]["Combined B centered"]),t)
            self.assertEqual(float(r[4]),float(shadow[t]["Prior A+B combined"]),t)
            self.assertEqual(float(r[5]),float(r[4]),t)   # GP=0 effective == prior
    def test_block3_pins_dal_nyg_natively(self):
        r=self.rows[-1]
        self.assertEqual(r[:5],["-1.07","-1.9","-0.83","1.6","0.77"])
        self.assertEqual(r[5],"NYG -0.8")
        self.assertEqual(r[7],"3.27")
        self.assertTrue(r[8].startswith("NYG"))
        self.assertEqual([r[9],r[10]],["22","24"])
    def test_probe_agrees_with_the_candidate_slate(self):
        g={r["GameID"]:r for r in csv.DictReader(
            open("audit/TTW_NFL_2026_Candidate_Week1_Slate_20260901.csv"))}["2026_01_DAL_NYG"]
        p=self.rows[-1]
        self.assertEqual((g["FinalMargin"],g["Model spread (fair line)"],g["SpreadEdge"]),
                         (p[4],p[5],p[7]))
    def test_no_unexpected_spreadsheet_errors_in_probe(self):
        blob="\n".join(",".join(r) for r in self.rows)
        for e in ("#REF!","#VALUE!","#DIV/0!","#N/A","#CIRC"):
            self.assertNotIn(e,blob,e)

class VerifierHasNoEscapePaths(unittest.TestCase):
    def test_no_or_true_escapes_remain(self):
        src=open("scripts/verify_srcb_blendfix_candidate.py").read()
        self.assertNotIn("or True",src,"verifier must not contain always-true escape paths")
    def test_settings_and_weight_checks_assert_real_values(self):
        src=open("scripts/verify_srcb_blendfix_candidate.py").read()
        for token in ("Current season","Current week (the week you are projecting)",
                      "As-of date (update each session; drives staleness checks)",
                      "Enable BET labels","ATS BET at >=","ATS INVESTIGATE at >=","ATS LEAN at >=",
                      "(0.4, 0.35, 0.25)"):
            self.assertIn(token,src,token)

class ProductionPostPromotion(unittest.TestCase):
    """Phase 5C post-promotion: production is live at v1.4. Verification was read-only;
    these tests pin the evidence so it cannot silently drift."""
    POST="TTW_NFL_2026_PROD_POSTPROMOTION_20260901T1650Z.xlsx"
    POST_SHA="fde0164c554283118ac6b14a6e765c5abc474add79ad845067cc94c48cb92da8"
    def test_post_promotion_export_pinned(self):
        self.assertEqual(hashlib.sha256(open(self.POST,'rb').read()).hexdigest(),self.POST_SHA)
    def test_manifest_readback_suite_passes(self):
        r=sh(f"python3 scripts/verify_production_post_promotion.py {self.POST}")
        self.assertEqual(r.returncode,0,r.stdout+r.stderr)
        self.assertIn("0 failed",r.stdout)
    def test_native_import_suite_passes_on_production(self):
        r=sh(f"python3 scripts/verify_native_import_5b.py {self.POST}")
        self.assertEqual(r.returncode,0,r.stdout+r.stderr)
        self.assertIn("0 failed",r.stdout)
        # that run rewrites the 5B artifact from production; restore it from the test copy
        sh("python3 scripts/verify_native_import_5b.py "
           "/tmp/claude-0/-home-user-TTW-nfl-2026/b271fc7c-3511-50b1-bc77-fb5e97203f27/"
           "scratchpad/testcopy_5b.xlsx")
    def test_production_is_live_at_v14(self):
        import openpyxl
        wb=openpyxl.load_workbook(self.POST,data_only=True)
        self.assertEqual(wb["START HERE"]["A1"].value,
                         "TO THE WINDOW \u2014 NFL POWER RATINGS 2026 (v1.4)")
        tr=wb["TEAM RATINGS"]
        self.assertTrue(all(tr.cell(r,4).value in (None,"") for r in range(5,37)),
                        "D column must be blank in production, not 0")
        i={tr.cell(r,1).value:r for r in range(5,37)}
        self.assertEqual((tr.cell(i["DAL"],10).value,tr.cell(i["DAL"],11).value),(-1.07,22))
        self.assertEqual((tr.cell(i["NYG"],10).value,tr.cell(i["NYG"],11).value),(-1.9,24))
        wb.close()

def load_tests(loader,tests,pattern):
    sys.path.insert(0,os.path.join(ROOT,"tests"))
    import test_blend_fix_semantics as B, test_gate_mutations as M
    tests.addTests(loader.loadTestsFromModule(B))
    tests.addTests(loader.loadTestsFromModule(M))
    return tests

if __name__=="__main__":
    unittest.main(verbosity=2)
