#!/usr/bin/env python3
"""Emit the v1.4 promotion manifest: every cell/range intended to change in the live
Google Sheet, with before and after values, read directly from the base export and the
finalized candidate.

Writes audit/TTW_NFL_2026_Promotion_Manifest_v14_20260901.csv

Read-only with respect to both workbooks. Nothing here promotes anything.
"""
import csv
import hashlib
import os

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BASE = os.path.join(ROOT, "TTW_NFL_2026_LIVE_EXPORT_20260901_BASE.xlsx")
CAND = os.path.join(ROOT, "TTW_NFL_Power_Ratings_2026_v1.4_SRCB_BLENDFIX_CANDIDATE.xlsx")
OUT = os.path.join(ROOT, "audit", "TTW_NFL_2026_Promotion_Manifest_v14_20260901.csv")

D_OLD = 'IFERROR(ROUND(INDEX(CalcGoverned,MATCH($A{r},CalcTeams,0)),2),"")'
D_NEW = ('IFERROR(IF(ISNUMBER(INDEX(CalcGoverned,MATCH($A{r},CalcTeams,0))),'
         'ROUND(INDEX(CalcGoverned,MATCH($A{r},CalcTeams,0)),2),""),"")')


def s(v):
    if v is None:
        return "(blank)"
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v)


def main():
    wb_b = openpyxl.load_workbook(BASE, data_only=True)
    wb_c = openpyxl.load_workbook(CAND, data_only=True)
    psb, psc = wb_b["PRESEASON"], wb_c["PRESEASON"]
    trb, trc = wb_b["TEAM RATINGS"], wb_c["TEAM RATINGS"]
    rows = []

    def add(sheet, cell, team, kind, before, after, note):
        rows.append([sheet, cell, team, kind, before, after, note])

    # 1 — Source B inputs (the only cells a human types)
    for i in range(32):
        r, t = 5 + i, psc.cell(5 + i, 2).value
        add("PRESEASON", "I%d" % r, t, "PASTE VALUE", s(psb.cell(r, 9).value),
            s(psc.cell(r, 9).value), "SrcB public avg — equal-weight VSiN p29 / ESPN FPI composite")
    for i in range(32):
        r, t = 5 + i, psc.cell(5 + i, 2).value
        add("PRESEASON", "K%d" % r, t, "PASTE TEXT", s(psb.cell(r, 11).value),
            (psc.cell(r, 11).value or "")[:90] + "…", "SrcB source citation")
    for i in range(32):
        r, t = 5 + i, psc.cell(5 + i, 2).value
        add("PRESEASON", "L%d" % r, t, "PASTE TEXT", s(psb.cell(r, 12).value),
            s(psc.cell(r, 12).value), "SrcB as-of date")

    # 2 — the one formula change
    for i in range(32):
        r, t = 5 + i, trc.cell(5 + i, 1).value
        add("TEAM RATINGS", "D%d" % r, t, "FORMULA", "=" + D_OLD.format(r=r),
            "=" + D_NEW.format(r=r), "ISNUMBER lookup protection (blank stays blank)")

    # 3 — banner and CHANGELOG
    add("START HERE", "A1", "", "TEXT", s(wb_b["START HERE"]["A1"].value),
        s(wb_c["START HERE"]["A1"].value), "Version banner v1.1 -> v1.4")
    chb, chc = wb_b["CHANGELOG"], wb_c["CHANGELOG"]
    for col, ci, label in (("A", 1, "Version"), ("B", 2, "Date"),
                           ("C", 3, "Change"), ("D", 4, "Backtest impact")):
        after = s(chc.cell(7, ci).value)
        add("CHANGELOG", "%s7" % col, "", "TEXT", s(chb.cell(7, ci).value),
            after if len(after) <= 90 else after[:90] + "…", "New v1.4 entry — %s" % label)

    # 4 — recalculated outputs the owner will SEE change but must NOT type
    for i in range(32):
        r, t = 5 + i, psc.cell(5 + i, 2).value
        for col, ci, note in (("J", 10, "SrcB centered"), ("S", 19, "Effective prior"),
                              ("T", 20, "Sources used")):
            add("PRESEASON", "%s%d" % (col, r), t, "RECALCULATED — DO NOT TYPE",
                s(psb.cell(r, ci).value), s(psc.cell(r, ci).value), note)
    for i in range(32):
        r, t = 5 + i, trc.cell(5 + i, 1).value
        for col, ci, note in (("F", 6, "Preseason prior"), ("H", 8, "Blended base"),
                              ("J", 10, "EFFECTIVE RATING"), ("K", 11, "Rank")):
            add("TEAM RATINGS", "%s%d" % (col, r), t, "RECALCULATED — DO NOT TYPE",
                s(trb.cell(r, ci).value), s(trc.cell(r, ci).value), note)
    add("TEAM RATINGS", "D5:D36", "all", "RECALCULATED — DO NOT TYPE", "0 (coerced)",
        "(blank)", "In-season governed rating at GP=0")
    add("ENGINE / DASHBOARD", "all Week-1 rows", "all", "RECALCULATED — DO NOT TYPE",
        "see slate CSV 'Baseline edge'", "see slate CSV 'SpreadEdge'",
        "16 games; recomputed from the new effective ratings")

    with open(OUT, "w", newline="") as f:
        w = csv.writer(f, lineterminator="\n")
        w.writerow(["Sheet", "Cell / Range", "Team", "Action", "Before", "After", "Note"])
        w.writerows(rows)

    typed = sum(1 for r in rows if r[3].startswith("PASTE") or r[3] == "FORMULA" or r[3] == "TEXT")
    print("wrote", OUT)
    print("rows: %d  (cells a human enters: %d, recalculated: %d)"
          % (len(rows), typed, len(rows) - typed))
    print("candidate sha256:", hashlib.sha256(open(CAND, "rb").read()).hexdigest())


if __name__ == "__main__":
    main()
