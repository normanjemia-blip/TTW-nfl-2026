#!/usr/bin/env python3
"""Produce a reproducible grounding JSON for the NFL workbook baseline audit.
Read-only. Does not modify the workbook."""
import hashlib, os, json, datetime
import openpyxl
from collections import Counter, defaultdict

WB="TTW_NFL_v1_1_1 Version 2.xlsx"
ERRS={"#REF!","#DIV/0!","#N/A","#VALUE!","#NAME?","#NULL!","#NUM!","#SPILL!","#CALC!"}

def sha256(p):
    h=hashlib.sha256()
    with open(p,"rb") as f:
        for ch in iter(lambda:f.read(1<<20),b""): h.update(ch)
    return h.hexdigest()

def is_input(v):
    if v is None: return False
    if isinstance(v,str):
        if v.strip()=="" or v.startswith("="): return False
    return True

g={"generated_utc":datetime.datetime.utcnow().isoformat()+"Z","workbook":WB}
g["size_bytes"]=os.path.getsize(WB)
g["sha256"]=sha256(WB)

wf=openpyxl.load_workbook(WB,data_only=False)
wv=openpyxl.load_workbook(WB,data_only=True)

g["sheet_count"]=len(wf.sheetnames)
g["sheet_order"]=wf.sheetnames
g["sheet_states"]={ws.title:ws.sheet_state for ws in wf.worksheets}
g["hidden_sheets"]=[t for t,s in g["sheet_states"].items() if s!="visible"]
g["visible_sheets"]=[t for t,s in g["sheet_states"].items() if s=="visible"]

# formulas
fc={}; tot=0
for ws in wf.worksheets:
    cnt=sum(1 for row in ws.iter_rows() for c in row
            if (isinstance(c.value,str) and c.value.startswith("=")) or c.data_type=="f")
    fc[ws.title]=cnt; tot+=cnt
g["formula_counts_by_sheet"]=fc
g["total_formula_cells"]=tot

# error cells (cached)
ec={}; etot=0; etype=Counter()
for ws in wv.worksheets:
    cnt=0
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value,str) and c.value in ERRS:
                cnt+=1; etot+=1; etype[c.value]+=1
    if cnt: ec[ws.title]=cnt
g["error_cells_total"]=etot
g["error_cells_by_sheet"]=ec
g["error_cells_by_type"]=dict(etype)

# version banner
g["version_banner_START_HERE_A1"]=wv["START HERE"].cell(1,1).value
chg=wf["CHANGELOG"]
g["changelog_versions"]=[chg.cell(r,1).value for r in range(2,chg.max_row+1) if chg.cell(r,1).value]

# schedule
ws=wf["IMPORT SCHEDULE"]
by_season=Counter(); by_st=Counter(); scored=Counter(); unscored=Counter()
for r in range(6,ws.max_row+1):
    gid=ws.cell(r,1).value; season=ws.cell(r,2).value; gt=ws.cell(r,3).value
    if gid is None or season is None: continue
    by_season[int(season)]+=1; by_st[f"{int(season)}_{gt}"]+=1
    if ws.cell(r,9).value is not None and ws.cell(r,11).value is not None: scored[int(season)]+=1
    else: unscored[int(season)]+=1
g["schedule_total_rows"]=sum(by_season.values())
g["schedule_by_season"]=dict(by_season)
g["schedule_by_season_gametype"]=dict(by_st)
g["schedule_scored_by_season"]=dict(scored)
g["schedule_unscored_by_season"]=dict(unscored)
g["schedule_2026_REG_count"]=by_st.get("2026_REG",0)

# production-state input counts
def count_inputs(sheet, cols, r0):
    ws=wf[sheet]; n=0
    for r in range(r0,ws.max_row+1):
        for c in cols:
            if is_input(ws.cell(r,c).value): n+=1
    return n
# market lines: current/upcoming actionable entries = rows with a usable market home spread (col Q non-blank) computed
wvm=wv["MARKET LINES"]; wfm=wf["MARKET LINES"]
usable_lines=0; sample_line_cells=0; sample_rows=0
for r in range(5,wfm.max_row+1):
    q=wvm.cell(r,17).value  # computed market home spread
    if q not in (None,""): usable_lines+=1
    row_inputs=[wfm.cell(r,c).value for c in (7,8,9,10,11)]
    ninp=sum(1 for v in row_inputs if is_input(v))
    if ninp:
        sample_line_cells+=ninp
        src=wfm.cell(r,14).value
        if isinstance(src,str) and "sample" in src.lower(): sample_rows+=1
g["market_lines_usable_market_spread_cells"]=usable_lines
g["market_lines_manual_input_cells_total"]=sample_line_cells
g["market_lines_sample_rows"]=sample_rows
g["adjustments_manual_input_cells"]=count_inputs("ADJUSTMENTS",(1,2,3,4,7,8,9,10,11,12),5)

# QB deltas nonzero
wsq=wf["QB VALUES"]; qb_nonzero=0; qb_rows=0
for r in range(5,wsq.max_row+1):
    if not is_input(wsq.cell(r,1).value): continue
    qb_rows+=1
    b=wsq.cell(r,3).value; a=wsq.cell(r,5).value
    if isinstance(b,(int,float)) and isinstance(a,(int,float)) and b!=a: qb_nonzero+=1
g["qb_rows"]=qb_rows
g["qb_nonzero_delta_rows"]=qb_nonzero
# team overrides
g["team_ratings_manual_override_cells"]=count_inputs("TEAM RATINGS",(9,),5)

# DQ summary (computed)
wd=wv["DATA QUALITY"]
g["dq_games_blocked_current_week"]=wd.cell(5,2).value
g["dq_games_warning_current_week"]=wd.cell(6,2).value
g["dq_unmatched_team_names"]=wd.cell(7,2).value

# settings key values (read-only, for the record — NOT changed)
wss=wv["SETTINGS"]
g["settings_current_season"]=wss.cell(5,2).value
g["settings_current_week"]=wss.cell(6,2).value
g["settings_asof_date"]=str(wss.cell(7,2).value)
g["settings_winmode"]=wss.cell(40,2).value

with open("audit/grounding.json","w") as f:
    json.dump(g,f,indent=2,default=str)
print(json.dumps(g,indent=2,default=str))
