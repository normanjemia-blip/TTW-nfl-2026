#!/usr/bin/env python3
import openpyxl
WB="TTW_NFL_Power_Ratings_2026_v1.1_AUTHORITATIVE.xlsx"
wf=openpyxl.load_workbook(WB, data_only=False)
ws=wf["QB VALUES"]
print("dims:", ws.dimensions, "max_row", ws.max_row, "max_col", ws.max_column)
print("\n=== Header rows 1-4 (all cols) ===")
for r in range(1,5):
    for c in range(1,15):
        v=ws.cell(r,c).value
        if v is not None:
            print(f"  {ws.cell(r,c).coordinate} = {repr(v)[:90]}")
    print("  ---")
print("\n=== Data rows 5-36: A-L (formula vs literal) ===")
for r in range(5,37):
    row=[]
    for c in range(1,13):
        cell=ws.cell(r,c)
        v=cell.value
        if v is None: continue
        tag="f" if (isinstance(v,str) and v.startswith("=")) else "L"
        row.append(f"{cell.coordinate}[{tag}]={repr(v)[:38]}")
    print(" | ".join(row))
