#!/usr/bin/env python3
import openpyxl, zipfile, hashlib, json, csv, sys
sys.path.insert(0,"scripts")
from market_dataset import WEEK1
BASE="TTW_NFL_Power_Ratings_2026_v1.2.1_QB_CANDIDATE.xlsx"
CAND="TTW_NFL_Power_Ratings_2026_v1.3_MARKET_CANDIDATE.xlsx"
AUTH="TTW_NFL_Power_Ratings_2026_v1.1_AUTHORITATIVE.xlsx"
def sha(p): return hashlib.sha256(open(p,"rb").read()).hexdigest()
rep={"base_sha256":sha(BASE),"candidate_sha256":sha(CAND),"authoritative_sha256":sha(AUTH)}

a=openpyxl.load_workbook(BASE,data_only=False); b=openpyxl.load_workbook(CAND,data_only=False)
rep["sheet_count"]=len(b.sheetnames)
rep["sheet_order_identical"]=a.sheetnames==b.sheetnames
rep["visibility_identical"]=[w.sheet_state for w in a.worksheets]==[w.sheet_state for w in b.worksheets]

def fmap(wb):
    m={}
    for sh in wb.worksheets:
        for row in sh.iter_rows():
            for c in row:
                v=c.value
                if (isinstance(v,str) and v.startswith("=")) or c.data_type=="f":
                    m[(sh.title,c.coordinate)]=v.text if hasattr(v,"text") else v
    return m
fa,fb=fmap(a),fmap(b)
rep["formula_count"]=len(fb); rep["formula_count_is_57399"]=len(fb)==57399
rep["formula_coords_identical"]=set(fa)==set(fb)
rep["formula_text_diffs"]=sum(1 for k in fa if fa[k]!=fb.get(k))

diffs=[]
for sh in a.sheetnames:
    x=a[sh]; y=b[sh]
    mr=max(x.max_row,y.max_row); mc=max(x.max_column,y.max_column)
    for r in range(1,mr+1):
        for c in range(1,mc+1):
            va=x.cell(r,c).value; vb=y.cell(r,c).value
            va=va.text if hasattr(va,"text") else va
            vb=vb.text if hasattr(vb,"text") else vb
            if va!=vb:
                diffs.append({"sheet":sh,"cell":x.cell(r,c).coordinate,
                              "old":("" if va is None else str(va)),"new":("" if vb is None else str(vb))})
rep["total_cell_diffs"]=len(diffs)
rep["diff_sheets"]=sorted(set(d["sheet"] for d in diffs))
rep["only_allowed_sheets_changed"]=set(rep["diff_sheets"])<={"MARKET LINES","START HERE","CHANGELOG"}
# phase-2 prohibitions
for name,sheet in [("qb_values_changes","QB VALUES"),("adjustment_changes","ADJUSTMENTS"),
                   ("team_rating_changes","TEAM RATINGS"),("settings_changes","SETTINGS"),
                   ("schedule_changes","IMPORT SCHEDULE"),("history_changes","HISTORY 2025"),
                   ("backtest_changes","BACKTEST")]:
    rep[name]=[d for d in diffs if d["sheet"]==sheet]
# market diffs must be confined to rows 5-20, cols G,H,I,N,O,P
import re as _re
bad=[]
for d in diffs:
    if d["sheet"]!="MARKET LINES": continue
    m=_re.match(r'([A-Z]+)(\d+)$', d["cell"])
    col,row=m.group(1),int(m.group(2))
    if not (5<=row<=20 and col in {"G","H","I","N","O","P"}): bad.append(d)
rep["market_diffs_outside_week1_inputs"]=bad
rep["market_cell_diffs"]=sum(1 for d in diffs if d["sheet"]=="MARKET LINES")
# sample rows 261-276 untouched
rep["sample_rows_untouched"]=not any(d["sheet"]=="MARKET LINES" and 261<=int(_re.match(r'[A-Z]+(\d+)$',d["cell"]).group(1))<=276 for d in diffs)

# Input-check validation (replicate sheet formula R): favorite must be in game, spread positive
mlv=openpyxl.load_workbook(CAND,data_only=True)["MARKET LINES"]
ml =b["MARKET LINES"]
checks=[]
for (row,away,home,fav,spread,total) in WEEK1:
    g=ml.cell(row,7).value; h=ml.cell(row,8).value; i=ml.cell(row,9).value
    cached_away=mlv.cell(row,5).value; cached_home=mlv.cell(row,6).value
    ok = (g in (cached_away,cached_home)) and isinstance(h,(int,float)) and h>0 and isinstance(i,(int,float))
    # expected market home spread: -h if fav==home else +h
    exp_q = -h if g==cached_home else h
    checks.append({"row":row,"game":f"{cached_away}@{cached_home}","favorite":g,"spread":h,"total":i,
                   "input_check":"OK" if ok else "FAIL","expected_market_home_spread":exp_q})
rep["input_checks"]=checks
rep["all_input_checks_ok"]=all(c["input_check"]=="OK" for c in checks)
rep["week1_rows_populated"]=sum(1 for c in checks if c["spread"] is not None)
rep["expected_usable_market_spreads"]=len([c for c in checks if c["favorite"] and c["spread"]])
rep["all_rows_expected_STALE"]=True  # line date 2026-05-15 vs AsOfDate 2026-07-13, StaleDays 3

zs=zipfile.ZipFile(BASE); zd=zipfile.ZipFile(CAND)
rep["zip_members_changed"]=[n for n in zs.namelist() if zs.read(n)!=zd.read(n)]
rep["drawings_persons_identical"]=all(zs.read(n)==zd.read(n) for n in zs.namelist() if "drawing" in n or "person" in n)

json.dump(rep,open("audit/market_candidate_verification.json","w"),indent=2,default=str)
with open("audit/market_changed_cells.csv","w",newline="") as f:
    w=csv.DictWriter(f,fieldnames=["sheet","cell","old","new"]); w.writeheader()
    for d in diffs: w.writerow({k:(v[:250] if isinstance(v,str) else v) for k,v in d.items()})
json.dump(diffs,open("audit/market_changed_cells.json","w"),indent=2,default=str)

for k in ["candidate_sha256","authoritative_sha256","sheet_count","sheet_order_identical","visibility_identical",
          "formula_count","formula_count_is_57399","formula_coords_identical","formula_text_diffs",
          "total_cell_diffs","diff_sheets","only_allowed_sheets_changed","market_cell_diffs",
          "market_diffs_outside_week1_inputs","sample_rows_untouched","qb_values_changes","adjustment_changes",
          "team_rating_changes","settings_changes","schedule_changes","history_changes","backtest_changes",
          "all_input_checks_ok","week1_rows_populated","zip_members_changed","drawings_persons_identical"]:
    print(f"{k}: {rep[k]}")
print("\n--- Week 1 input checks ---")
for c in checks: print(f"  r{c['row']} {c['game']:9s} fav={c['favorite']:4s} spr={c['spread']:5} tot={c['total']:5} -> {c['input_check']}  (mkt home spread {c['expected_market_home_spread']})")
