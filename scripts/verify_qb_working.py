#!/usr/bin/env python3
import openpyxl, zipfile, hashlib, json
SRC="TTW_NFL_Power_Ratings_2026_v1.1_AUTHORITATIVE.xlsx"
DST="TTW_NFL_Power_Ratings_2026_v1.2_QB_WORKING.xlsx"

def sha(p): return hashlib.sha256(open(p,"rb").read()).hexdigest()
rep={}
rep["source_sha256"]=sha(SRC); rep["working_sha256"]=sha(DST)

ws_=openpyxl.load_workbook(SRC,data_only=False)
wd_=openpyxl.load_workbook(DST,data_only=False)
rep["sheet_count_working"]=len(wd_.sheetnames)
rep["sheet_order_identical"]=ws_.sheetnames==wd_.sheetnames
rep["visibility_identical"]=[w.sheet_state for w in ws_.worksheets]==[w.sheet_state for w in wd_.worksheets]

def fmap(wb):
    m={}
    for sh in wb.worksheets:
        for row in sh.iter_rows():
            for c in row:
                v=c.value
                if (isinstance(v,str) and v.startswith("=")) or c.data_type=="f":
                    m[(sh.title,c.coordinate)]=v.text if hasattr(v,"text") else v
    return m
fs,fd=fmap(ws_),fmap(wd_)
rep["formula_count_source"]=len(fs); rep["formula_count_working"]=len(fd)
rep["formula_coords_identical"]=set(fs)==set(fd)
rep["formula_text_diffs"]=sum(1 for k in fs if fs[k]!=fd.get(k))

# full cell diff
diffs=[]
for sh in ws_.sheetnames:
    a=ws_[sh]; b=wd_[sh]
    mr=max(a.max_row,b.max_row); mc=max(a.max_column,b.max_column)
    for r in range(1,mr+1):
        for c in range(1,mc+1):
            va=a.cell(r,c).value; vb=b.cell(r,c).value
            va=va.text if hasattr(va,"text") else va
            vb=vb.text if hasattr(vb,"text") else vb
            if va!=vb:
                diffs.append({"sheet":sh,"cell":a.cell(r,c).coordinate,
                              "old":(str(va)[:45] if va is not None else None),
                              "new":(str(vb)[:60] if vb is not None else None)})
rep["total_cell_diffs"]=len(diffs)
rep["nonQB_cell_diffs"]=[d for d in diffs if d["sheet"]!="QB VALUES"]
rep["qb_cell_diff_count"]=sum(1 for d in diffs if d["sheet"]=="QB VALUES")

# zip member parity
zs=zipfile.ZipFile(SRC); zd=zipfile.ZipFile(DST)
rep["zip_members_source"]=len(zs.namelist()); rep["zip_members_working"]=len(zd.namelist())
changed=[n for n in zs.namelist() if zs.read(n)!=zd.read(n)]
rep["zip_members_changed"]=changed
rep["drawings_persons_identical"]=all(zs.read(n)==zd.read(n) for n in zs.namelist() if "drawing" in n or "person" in n)

# QB delta check (F,G,H must all be 0 -> output neutral)
wdv=openpyxl.load_workbook(DST,data_only=True)  # cached from source (formulas not recalced by us)
# Instead verify E-C logic: for each row, delta should be 0 (either 0-0 or blank)
q=wd_["QB VALUES"]
nonzero_delta=[]
for r in range(5,37):
    C=q.cell(r,3).value; E=q.cell(r,5).value
    # delta = 0 if E or C blank, else E-C
    if C not in (None,"") and E not in (None,""):
        if (E-C)!=0: nonzero_delta.append((q.cell(r,1).value, E-C))
rep["qb_rows_with_nonzero_delta"]=nonzero_delta

# production-state style counts on working
def cnt_conf(val):
    return sum(1 for r in range(5,37) if q.cell(r,9).value==val)
rep["confidence_High"]=cnt_conf("High"); rep["confidence_Low"]=cnt_conf("Low"); rep["confidence_Medium"]=cnt_conf("Medium")
rep["baseline_value_zero"]=sum(1 for r in range(5,37) if q.cell(r,3).value==0)
rep["baseline_value_blank"]=sum(1 for r in range(5,37) if q.cell(r,3).value in (None,""))

json.dump(rep, open("audit/qb_working_verification.json","w"), indent=2, default=str)
for k in ["source_sha256","working_sha256","sheet_count_working","sheet_order_identical","visibility_identical",
          "formula_count_source","formula_count_working","formula_coords_identical","formula_text_diffs",
          "total_cell_diffs","qb_cell_diff_count","zip_members_changed","drawings_persons_identical",
          "qb_rows_with_nonzero_delta","confidence_High","confidence_Low","confidence_Medium",
          "baseline_value_zero","baseline_value_blank"]:
    print(f"{k}: {rep[k]}")
print("\nNON-QB cell diffs (must be empty):", rep["nonQB_cell_diffs"])
