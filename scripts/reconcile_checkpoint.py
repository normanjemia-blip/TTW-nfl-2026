#!/usr/bin/env python3
"""Reconcile the stated 2026-08-15 checkpoint against the AUTHORITATIVE workbook. Read-only."""
import openpyxl, datetime
A="TTW_NFL_Power_Ratings_2026_v1.1_AUTHORITATIVE.xlsx"
wf=openpyxl.load_workbook(A,data_only=False); wv=openpyxl.load_workbook(A,data_only=True)
out={}
s=wv["SETTINGS"]
out["as_of_date"]=str(s.cell(7,2).value); out["current_season"]=s.cell(5,2).value; out["current_week"]=s.cell(6,2).value
out["EnableBetLabels(B67)"]=s.cell(67,2).value
out["thresholds ATS BET/INV/LEAN"]=[s.cell(26,2).value,s.cell(27,2).value,s.cell(28,2).value]
out["thresholds TOT BET/INV/LEAN"]=[s.cell(29,2).value,s.cell(30,2).value,s.cell(31,2).value]
out["win_totals_mode(B40)"]=s.cell(40,2).value
# 272 REG 2026
sch=wf["IMPORT SCHEDULE"]; reg=0
for r in range(6,sch.max_row+1):
    if sch.cell(r,2).value==2026 and sch.cell(r,3).value=="REG": reg+=1
out["2026_REG_games"]=reg
# Week1 market completeness in AUTHORITATIVE
mv=wv["MARKET LINES"]; mf=wf["MARKET LINES"]
miss_s=miss_t=0
for r in range(5,21):
    if mf.cell(r,8).value in (None,""): miss_s+=1
    if mf.cell(r,9).value in (None,""): miss_t+=1
out["wk1_missing_spreads"]=miss_s; out["wk1_missing_totals"]=miss_t
out["usable_market_spreads"]=sum(1 for r in range(5,mv.max_row+1) if mv.cell(r,17).value not in (None,""))
# QB VALUES populated in AUTHORITATIVE?
q=wf["QB VALUES"]; nonblank=sum(1 for r in range(5,37) if q.cell(r,3).value not in (None,""))
zeroed=sum(1 for r in range(5,37) if q.cell(r,3).value==0)
out["QB rows with baseline value"]=nonblank; out["QB rows zero-initialized"]=zeroed
out["QB last_update sample(ARI)"]=str(q.cell(5,11).value)
# PRESEASON sources
p=wf["PRESEASON"]
out["PRESEASON header D/I/other"]=[p.cell(4,c).value for c in range(4,12)]
srcB=sum(1 for r in range(5,37) if p.cell(r,9).value not in (None,""))
out["PRESEASON SrcB(public) populated rows"]=srcB
srcA=sum(1 for r in range(5,37) if p.cell(r,4).value not in (None,""))
out["PRESEASON SrcA(TTW regressed) populated rows"]=srcA
# DATA QUALITY div0
dq=wv["DATA QUALITY"]
out["DQ_B8 (rating mean check)"]=dq.cell(8,2).value
out["DQ blocked/warn/unmatched"]=[dq.cell(5,2).value,dq.cell(6,2).value,dq.cell(7,2).value]
out["sheets"]=len(wf.sheetnames)
for k,v in out.items(): print(f"{k}: {v}")
