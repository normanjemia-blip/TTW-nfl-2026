#!/usr/bin/env python3
import openpyxl
WB="TTW_NFL_v1_1_1 Version 2.xlsx"
wb=openpyxl.load_workbook(WB, data_only=False)

def dump(sheet, maxr=40, maxc=8):
    ws=wb[sheet]
    print(f"\n===== {sheet} (state={ws.sheet_state}) =====")
    for r in range(1, min(maxr, ws.max_row)+1):
        cells=[]
        for c in range(1, min(maxc, ws.max_column)+1):
            v=ws.cell(r,c).value
            if v is not None:
                cells.append(f"{ws.cell(r,c).coordinate}={repr(v)[:60]}")
        if cells:
            print(" | ".join(cells))

for s in ["START HERE","CHANGELOG","DICTIONARY"]:
    dump(s, 40, 4)
