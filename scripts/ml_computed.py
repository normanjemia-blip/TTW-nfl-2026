#!/usr/bin/env python3
import openpyxl
WB="TTW_NFL_v1_1_1 Version 2.xlsx"
wbv=openpyxl.load_workbook(WB, data_only=True)
ws=wbv["MARKET LINES"]
print("Cached computed values. Cols: A Seq, B GameID, C Week, E Away, F Home, G Fav, H Spread, I Total, N Source")
def show(rng):
    for r in rng:
        vals={ 'Seq':ws.cell(r,1).value,'GameID':ws.cell(r,2).value,'Week':ws.cell(r,3).value,
               'Away':ws.cell(r,5).value,'Home':ws.cell(r,6).value,'Fav':ws.cell(r,7).value,
               'Spr':ws.cell(r,8).value,'Tot':ws.cell(r,9).value,'Src':ws.cell(r,14).value}
        if any(v not in (None,"") for v in vals.values()):
            print(f"R{r}:", {k:v for k,v in vals.items() if v not in (None,"")})
print("\n--- Entry area rows 5-40 ---")
show(range(5,41))
print("\n--- Sample area rows 260-278 ---")
show(range(260,279))
