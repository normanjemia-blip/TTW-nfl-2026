#!/usr/bin/env python3
import openpyxl, zipfile, re
WB="TTW_NFL_Power_Ratings_2026_v1.1_AUTHORITATIVE.xlsx"
wf=openpyxl.load_workbook(WB, data_only=False)
ws=wf["QB VALUES"]
print("=== L5 (Check), N5 (QBFlag) formulas ===")
print("L5:", ws["L5"].value)
print("N5:", ws["N5"].value)
print("M5 (Reviewed season) current value:", repr(ws["M5"].value))
print("K5 (Last update) current value:", repr(ws["K5"].value))
# Does any formula reference M (Reviewed) col or QBConf/QBStaleDays?
print("\n=== references to QB VALUES col M / K / I / QBConf / QBStaleDays ===")
for wsx in wf.worksheets:
    for row in wsx.iter_rows():
        for c in row:
            v=c.value
            if isinstance(v,str) and v.startswith("="):
                if re.search(r"'QB VALUES'!\$?M|'QB VALUES'!\$?K|QBStaleDays|QBConf|ReviewedSeason", v):
                    print(f"  {wsx.title}!{c.coordinate}: {v[:110]}")
                    break
# Map sheet6 = QB VALUES? confirm via workbook rels
print("\n=== raw sheet6.xml rows 4-6 ===")
z=zipfile.ZipFile(WB)
xml=z.read("xl/worksheets/sheet6.xml").decode("utf-8")
for rr in (4,5,6,23):
    m=re.search(r'<row r="%d"[^>]*>.*?</row>'%rr, xml)
    print(f"\n--- ROW {rr} ---")
    print(m.group(0) if m else "not found")
