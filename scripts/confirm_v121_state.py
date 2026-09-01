#!/usr/bin/env python3
"""Confirm the v1.2.1 QB checkpoint state before the Aug-10 sweep. Read-only."""
import openpyxl, datetime
WB="TTW_NFL_Power_Ratings_2026_v1.2.1_QB_CANDIDATE.xlsx"
wf=openpyxl.load_workbook(WB,data_only=False)
q=wf["QB VALUES"]
CUR=2026; ASOF=datetime.datetime(2026,7,13); STALE=30

print("sheets:",len(wf.sheetnames))
n=0
for sh in wf.worksheets:
    for row in sh.iter_rows():
        for c in row:
            v=c.value
            if (isinstance(v,str) and v.startswith("=")) or c.data_type=="f": n+=1
print("formulas:",n,"-> 57,399?" , n==57399)

ok=unc=0; rows=[]; nonzero=[]
for r in range(5,37):
    team=q.cell(r,1).value; B=q.cell(r,2).value; C=q.cell(r,3).value
    D=q.cell(r,4).value; E=q.cell(r,5).value; I=q.cell(r,9).value
    K=q.cell(r,11).value; M=q.cell(r,13).value
    delta = 0 if (C in (None,"") or E in (None,"")) else round(E-C,4)
    if delta!=0: nonzero.append((team,B,D,delta))
    stale = isinstance(K,datetime.datetime) and (ASOF-K).days>STALE
    flag = 1 if (D in (None,"") or I=="Low" or M!=CUR or stale) else 0
    (globals().__setitem__('unc',unc+1) if flag else globals().__setitem__('ok',ok+1))
    ok = ok+ (0 if flag else 1) if False else ok
    rows.append((r,team,B,C,D,E,delta,I,"UNCERTAIN" if flag else "OK"))
ok=sum(1 for x in rows if x[8]=="OK"); unc=sum(1 for x in rows if x[8]=="UNCERTAIN")
print("teams:",len(rows))
print("QB OK:",ok," QB UNCERTAIN:",unc)
print("UNCERTAIN teams:",[x[1] for x in rows if x[8]=="UNCERTAIN"])
print("nonzero deviations:",nonzero)
print("\n--- baseline roster (team | baselineQB | activeQB | delta | conf | status) ---")
for (r,t,B,C,D,E,d,I,s) in rows:
    print(f"  r{r:<3}{t:<4}| {str(B):<18}| {str(D):<18}| {d:<5}| {str(I):<7}| {s}")
