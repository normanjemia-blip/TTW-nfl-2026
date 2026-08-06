#!/usr/bin/env python3
import openpyxl, zipfile, re
WB="TTW_NFL_Power_Ratings_2026_v1.2.1_QB_CANDIDATE.xlsx"
wf=openpyxl.load_workbook(WB,data_only=False)
wv=openpyxl.load_workbook(WB,data_only=True)
ws=wf["MARKET LINES"]; vs=wv["MARKET LINES"]

print("=== Header row 4 ===")
for c in range(1,20):
    v=ws.cell(4,c).value
    if v: print(f"  {ws.cell(4,c).coordinate} = {v}")

print("\n=== Instruction rows 1-3 ===")
for r in (1,2,3):
    v=ws.cell(r,1).value
    if v: print(f"  A{r}: {v}")

print("\n=== Week 1 rows (computed Week==1) ===")
wk1=[]
for r in range(5,60):
    wk=vs.cell(r,3).value
    if wk==1:
        wk1.append(r)
        print(f"  row {r}: GameID={vs.cell(r,2).value!r} Away={vs.cell(r,5).value!r} Home={vs.cell(r,6).value!r} "
              f"Date={vs.cell(r,4).value!r} Fav={vs.cell(r,7).value!r} Spread={vs.cell(r,8).value!r} Total={vs.cell(r,9).value!r}")
print("Week1 row count:", len(wk1), "rows", wk1[0], "-", wk1[-1])

print("\n=== raw XML of MARKET LINES row 5 (cell forms) ===")
z=zipfile.ZipFile(WB)
# find sheet file for MARKET LINES -> sheet4.xml per earlier mapping
xml=z.read("xl/worksheets/sheet4.xml").decode("utf-8")
m=re.search(r'<row r="5"[^>]*>.*?</row>', xml)
print(m.group(0)[:1200])
