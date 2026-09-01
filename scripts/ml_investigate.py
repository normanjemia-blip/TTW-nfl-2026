#!/usr/bin/env python3
import openpyxl
WB="TTW_NFL_v1_1_1 Version 2.xlsx"
wb=openpyxl.load_workbook(WB, data_only=False)
ws=wb["MARKET LINES"]
print("max_row", ws.max_row, "max_col", ws.max_column)
# find all non-empty rows and their column A + literal content, scanning whole sheet
print("\n-- Rows with any literal (non-formula) value in cols G-K or a section label in A --")
for r in range(1, ws.max_row+1):
    a=ws.cell(r,1).value
    rowvals=[]
    for c in range(1, ws.max_column+1):
        v=ws.cell(r,c).value
        if v is not None and not (isinstance(v,str) and v.startswith("=")):
            rowvals.append(f"{ws.cell(r,c).coordinate}={repr(v)[:40]}")
    if rowvals and (r<=5 or r>=255 or (isinstance(a,str) and not a.replace('.','').isdigit())):
        print(f"R{r}:", " | ".join(rowvals))
