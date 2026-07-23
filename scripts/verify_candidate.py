#!/usr/bin/env python3
"""Verify candidate vs source: only proven version-label cells + new CHANGELOG row differ.
Everything else (sheets, formulas, coords, order, visibility, data) identical."""
import openpyxl, hashlib, json, zipfile

SRC="TTW_NFL_v1_1_1 Version 2.xlsx"
DST="TTW_NFL_Power_Ratings_2026_v1.1_VERSION_ALIGNMENT_CANDIDATE.xlsx"

def sha(p):
    h=hashlib.sha256(); h.update(open(p,"rb").read()); return h.hexdigest()

rep={"source":SRC,"candidate":DST}
rep["source_sha256"]=sha(SRC)
rep["candidate_sha256"]=sha(DST)
rep["source_size"]=__import__("os").path.getsize(SRC)
rep["candidate_size"]=__import__("os").path.getsize(DST)

# zip member parity
zs=set(zipfile.ZipFile(SRC).namelist()); zd=set(zipfile.ZipFile(DST).namelist())
rep["zip_members_source"]=len(zs); rep["zip_members_candidate"]=len(zd)
rep["zip_members_added"]=sorted(zd-zs); rep["zip_members_removed"]=sorted(zs-zd)

ws_=openpyxl.load_workbook(SRC,data_only=False)
wd_=openpyxl.load_workbook(DST,data_only=False)

rep["sheet_count_source"]=len(ws_.sheetnames)
rep["sheet_count_candidate"]=len(wd_.sheetnames)
rep["sheet_order_identical"]=ws_.sheetnames==wd_.sheetnames
rep["visibility_identical"]=([w.sheet_state for w in ws_.worksheets]==[w.sheet_state for w in wd_.worksheets])

def formula_map(wb):
    m={}
    tot=0
    for sh in wb.worksheets:
        for row in sh.iter_rows():
            for c in row:
                v=c.value
                if (isinstance(v,str) and v.startswith("=")) or c.data_type=="f":
                    m[(sh.title,c.coordinate)]= v.text if hasattr(v,"text") else v
                    tot+=1
    return m,tot

fs,ts=formula_map(ws_); fd,td=formula_map(wd_)
rep["formula_count_source"]=ts
rep["formula_count_candidate"]=td
rep["formula_coords_identical"]=set(fs.keys())==set(fd.keys())
diff_formulas=[k for k in fs if fs[k]!=fd.get(k)]
rep["formula_text_diffs"]=len(diff_formulas)
rep["formula_text_diff_examples"]=[{"cell":f"{k[0]}!{k[1]}"} for k in diff_formulas[:10]]

# full cell-value diff (all sheets, all cells)
diffs=[]
for sh in ws_.sheetnames:
    a=ws_[sh]; b=wd_[sh]
    maxr=max(a.max_row,b.max_row); maxc=max(a.max_column,b.max_column)
    for r in range(1,maxr+1):
        for c in range(1,maxc+1):
            va=a.cell(r,c).value; vb=b.cell(r,c).value
            # normalize array formula objects
            va_=va.text if hasattr(va,"text") else va
            vb_=vb.text if hasattr(vb,"text") else vb
            if va_!=vb_:
                diffs.append({"sheet":sh,"cell":a.cell(r,c).coordinate,
                              "source":(str(va_)[:60] if va_ is not None else None),
                              "candidate":(str(vb_)[:60] if vb_ is not None else None)})
rep["total_cell_diffs"]=len(diffs)
rep["cell_diffs"]=diffs

# schedule 2026 REG count parity
def sched_2026(wb):
    w=wb["IMPORT SCHEDULE"]; n=0
    for r in range(6,w.max_row+1):
        if w.cell(r,2).value==2026 and w.cell(r,3).value=="REG": n+=1
    return n
rep["sched_2026_REG_source"]=sched_2026(ws_)
rep["sched_2026_REG_candidate"]=sched_2026(wd_)

json.dump(rep, open("audit/candidate_verification.json","w"), indent=2, default=str)
# print summary
for k in ["source_sha256","candidate_sha256","source_size","candidate_size",
          "zip_members_source","zip_members_candidate","zip_members_added","zip_members_removed",
          "sheet_count_source","sheet_count_candidate","sheet_order_identical","visibility_identical",
          "formula_count_source","formula_count_candidate","formula_coords_identical","formula_text_diffs",
          "sched_2026_REG_source","sched_2026_REG_candidate","total_cell_diffs"]:
    print(f"{k}: {rep[k]}")
print("\nCELL DIFFS:")
for d in rep["cell_diffs"]:
    print(f"  [{d['sheet']}!{d['cell']}]  {d['source']!r}  ->  {d['candidate']!r}")
