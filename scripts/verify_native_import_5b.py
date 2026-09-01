#!/usr/bin/env python3
"""Phase 5B — read-only verification of the native Google Sheets import of the v1.4
candidate, against an xlsx export of that test copy.

Verifies the full promotion manifest, the 16-game ENGINE/DASHBOARD reconciliation, and
scans every tab for formula/conversion errors introduced by the native round trip.

Usage: verify_native_import_5b.py <testcopy_export.xlsx>

Read-only. Touches no production artifact and no Google Sheet.
"""
import csv
import os
import sys

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BASE = os.path.join(ROOT, "TTW_NFL_2026_LIVE_EXPORT_20260901_BASE.xlsx")
CAND = os.path.join(ROOT, "TTW_NFL_Power_Ratings_2026_v1.4_SRCB_BLENDFIX_CANDIDATE.xlsx")
SLATE = os.path.join(ROOT, "audit", "TTW_NFL_2026_Candidate_Week1_Slate_20260901.csv")
MANIFEST = os.path.join(ROOT, "audit", "TTW_NFL_2026_Promotion_Manifest_v14_20260901.csv")
OUT = os.path.join(ROOT, "audit", "TTW_NFL_2026_Native_Import_Reconciliation_20260901.csv")

BANNER = "TO THE WINDOW — NFL POWER RATINGS 2026 (v1.4)"
D_NEW = ('IFERROR(IF(ISNUMBER(INDEX(CalcGoverned,MATCH($A{r},CalcTeams,0))),'
         'ROUND(INDEX(CalcGoverned,MATCH($A{r},CalcTeams,0)),2),""),"")')
ERRORS = ("#REF!", "#VALUE!", "#N/A", "#DIV/0!", "#NAME?", "#NUM!", "#NULL!", "#ERROR!")
# Pre-existing in production today: empty-range AVERAGE at zero games played.
# Documented in audit/DQ_DIV0_Diagnosis.md; unrelated to the v1.4 candidate.
KNOWN_ERRORS = {("CALC", "B%d" % r) for r in range(39, 44)} | {("DATA QUALITY", "B8")}
TABS = ["START HERE", "DASHBOARD", "ENGINE", "MARKET LINES", "ADJUSTMENTS", "QB VALUES",
        "PRESEASON MONITOR", "TEAM RATINGS", "DATA QUALITY", "SETTINGS", "IMPORT SCHEDULE",
        "IMPORT STATS", "MAP", "CLEAN", "CALC", "LISTS", "PRESEASON", "HISTORY 2025",
        "BACKTEST", "AUDIT", "DICTIONARY", "CHANGELOG"]

R = []
def check(name, ok, detail=""):
    R.append((name, bool(ok), detail))


def ftext(cell):
    v = cell.value
    if hasattr(v, "text"):
        return v.text.lstrip("=")
    return v[1:] if isinstance(v, str) and v.startswith("=") else None


def fmt1(x):
    return ("%+.1f" % x).replace("+-", "-")


def main(path):
    nf = openpyxl.load_workbook(path)               # formulas
    nv = openpyxl.load_workbook(path, data_only=True)  # Google's computed values
    cv = openpyxl.load_workbook(CAND, data_only=True)
    bv = openpyxl.load_workbook(BASE, data_only=True)

    # ---------- structure ----------
    check("22_tabs_preserved", nf.sheetnames == TABS,
          "%d tabs%s" % (len(nf.sheetnames),
                         "" if nf.sheetnames == TABS else ": %s" % nf.sheetnames))
    nfc = sum(1 for s in nf.sheetnames for row in nf[s].iter_rows()
              for c in row if ftext(c) is not None)
    check("formula_count_57399", nfc == 57399, str(nfc))

    # ---------- manifest: the 133 cells a human enters ----------
    ps, tr = nv["PRESEASON"], nv["TEAM RATINGS"]
    psc, trc = cv["PRESEASON"], cv["TEAM RATINGS"]
    teams = [ps.cell(r, 2).value for r in range(5, 37)]
    bad = [t for i, t in enumerate(teams)
           if abs((ps.cell(5 + i, 9).value or 0) - psc.cell(5 + i, 9).value) > 1e-9]
    check("srcB_paste_I5_I36_intact", not bad, "mismatched: %s" % bad if bad else "32/32")
    check("srcB_source_K5_K36_intact",
          all((ps.cell(r, 11).value or "").startswith("Equal-weight composite")
              for r in range(5, 37)), "32/32")
    check("srcB_asof_L5_L36_intact",
          all(str(ps.cell(r, 12).value)[:10] == "2026-09-01" for r in range(5, 37)), "32/32")
    ntr = nf["TEAM RATINGS"]
    check("D5_D36_formula_survived_round_trip",
          all(ftext(ntr.cell(r, 4)) == D_NEW.format(r=r) for r in range(5, 37)),
          "ISNUMBER lookup protection present in all 32")
    check("banner_reads_v14", nv["START HERE"]["A1"].value == BANNER,
          repr(nv["START HERE"]["A1"].value))
    ch = nv["CHANGELOG"]
    check("changelog_v14_entry_present",
          ch.cell(7, 1).value == "1.4" and ch.cell(7, 2).value == "2026-09-01"
          and "ISNUMBER" in (ch.cell(7, 3).value or ""), "row 7 intact")

    # ---------- the behaviour that mattered ----------
    dblank = [teams[i] for i in range(32) if tr.cell(5 + i, 4).value not in (None, "")]
    check("D5_D36_genuinely_blank_not_zero", not dblank,
          "non-blank: %s" % dblank if dblank else "32/32 blank (0 would prove the fix failed)")
    check("GP_zero_all_32", all(tr.cell(r, 2).value == 0 for r in range(5, 37)))
    fhj = [teams[i] for i in range(32)
           if not (tr.cell(5 + i, 6).value == tr.cell(5 + i, 8).value == tr.cell(5 + i, 10).value)]
    check("F_equals_H_equals_J_all_32", not fhj, "mismatched: %s" % fhj if fhj else "32/32")
    pri = [teams[i] for i in range(32)
           if tr.cell(5 + i, 10).value != psc.cell(5 + i, 19).value]
    check("effective_rating_equals_candidate_prior", not pri, "mismatched: %s" % pri if pri else "32/32")
    rk = [teams[i] for i in range(32) if tr.cell(5 + i, 11).value != trc.cell(5 + i, 11).value]
    check("ranks_match_candidate", not rk, "mismatched: %s" % rk if rk else "32/32")
    check("sources_used_is_2_all_32", all(ps.cell(r, 20).value == 2 for r in range(5, 37)))
    jm = [teams[i] for i in range(32) if ps.cell(5 + i, 10).value != psc.cell(5 + i, 10).value]
    check("srcB_centered_matches_candidate", not jm, "mismatched: %s" % jm if jm else "32/32")
    sm = [teams[i] for i in range(32) if ps.cell(5 + i, 19).value != psc.cell(5 + i, 19).value]
    check("effective_prior_matches_candidate", not sm, "mismatched: %s" % sm if sm else "32/32")

    # ---------- guardrails ----------
    check("srcC_blank_all_32", all(ps.cell(r, 14).value is None for r in range(5, 37)))
    bw = [teams[i] for i in range(32)
          if (ps.cell(5 + i, 8).value, ps.cell(5 + i, 13).value, ps.cell(5 + i, 18).value)
          != (0.4, 0.35, 0.25)]
    check("weights_A040_B035_C025", not bw, "offending: %s" % bw if bw else "32/32")
    st = {str(nv["SETTINGS"].cell(r, 1).value).strip(): nv["SETTINGS"].cell(r, 2).value
          for r in range(1, 80) if nv["SETTINGS"].cell(r, 1).value is not None}
    want = [("Current season", 2026), ("Current week (the week you are projecting)", 1),
            ("As-of date (update each session; drives staleness checks)", "2026-09-01"),
            ("Enable BET labels (OFF = 1.5+ ATS edges display STRONG INVESTIGATE)", "Y"),
            ("ATS BET at >=", 1.5), ("ATS INVESTIGATE at >=", 1.5), ("ATS LEAN at >=", 1.0),
            ("Win-totals mode (VALIDATE-ONLY until conversion is verified)", "VALIDATE-ONLY"),
            ("Home-field advantage, points", 1.6),
            ("Prior-season regression to mean (0-1)", 0.33)]
    bs = []
    for k, v in want:
        g = st.get(k, "<missing>")
        if hasattr(g, "strftime"):
            g = g.strftime("%Y-%m-%d")
        elif isinstance(g, float) and isinstance(v, int) and g == int(g):
            g = int(g)
        if g != v:
            bs.append("%s: got %r want %r" % (k, g, v))
    check("settings_intact", not bs, "; ".join(bs) if bs else "10/10 values")
    check("blend_wt_wk1_080", nv["TEAM RATINGS"].cell(5, 7).value == 0.8)

    # ---------- untouched production inputs ----------
    for sheet, rng in (("MARKET LINES", (5, 61, 1, 19)), ("QB VALUES", (5, 36, 1, 14)),
                       ("ADJUSTMENTS", (5, 104, 1, 13)), ("PRESEASON MONITOR", (1, 40, 1, 18))):
        r0, r1, c0, c1 = rng
        diff = [(r, c) for r in range(r0, r1 + 1) for c in range(c0, c1 + 1)
                if bv[sheet].cell(r, c).value != nv[sheet].cell(r, c).value]
        check("unchanged_" + sheet.replace(" ", "_").lower(), not diff,
              "%d differing cells" % len(diff) if diff else "identical to the pre-change base")

    # ---------- error / conversion scan across every tab ----------
    found = []
    for s in nv.sheetnames:
        for row in nv[s].iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and any(e in v for e in ERRORS):
                    if (s, c.coordinate) not in KNOWN_ERRORS:
                        found.append((s, c.coordinate, v[:40]))
    check("no_new_formula_or_conversion_errors", not found,
          "%d unexpected: %s" % (len(found), found[:6]) if found
          else "only the 6 pre-existing #DIV/0! cells (CALC B39:B43, DATA QUALITY B8)")
    known_present = [k for k in sorted(KNOWN_ERRORS)
                     if isinstance(nv[k[0]][k[1]].value, str)
                     and "#DIV/0!" in str(nv[k[0]][k[1]].value)]
    check("preexisting_div0_unchanged_in_count", len(known_present) == 6,
          "%d/6 present (pre-existing, documented, not caused by v1.4)" % len(known_present))

    # ---------- 16-game reconciliation ----------
    slate = {r["GameID"]: r for r in csv.DictReader(open(SLATE))}
    en, da = nv["ENGINE"], nv["DASHBOARD"]
    rows, mis = [], []
    for r in range(5, 300):
        gid = en.cell(r, 2).value
        if not gid or en.cell(r, 3).value != 1:
            continue
        s = slate[gid]
        got = (round(en.cell(r, 19).value, 2), en.cell(r, 20).value,
               round(en.cell(r, 22).value, 2), en.cell(r, 24).value)
        exp = (float(s["FinalMargin"]), s["Model spread (fair line)"],
               float(s["SpreadEdge"]), s["Supported side"])
        ok = got[0] == exp[0] and got[1] == exp[1] and got[2] == exp[2] and got[3] == exp[3]
        if not ok:
            mis.append((gid, got, exp))
        rows.append([gid, en.cell(r, 5).value, en.cell(r, 6).value,
                     en.cell(r, 10).value, en.cell(r, 11).value, got[0], got[1],
                     en.cell(r, 21).value, got[2], got[3], en.cell(r, 26).value or "",
                     "MATCH" if ok else "MISMATCH"])
    check("engine_16_week1_rows", len(rows) == 16, str(len(rows)))
    check("engine_reconciles_with_slate_csv", not mis,
          "mismatches: %s" % mis[:3] if mis else "16/16 games match FinalMargin, fair line, edge and side")
    dash = sum(1 for r in range(1, 40)
               if isinstance(da.cell(r, 1).value, str) and " @ " in str(da.cell(r, 1).value))
    check("dashboard_shows_16_games", dash == 16, str(dash))
    dmis = []
    for r in range(1, 40):
        lbl = da.cell(r, 1).value
        if isinstance(lbl, str) and " @ " in lbl:
            away = lbl.split(" @ ")[0].strip()
            row = [x for x in rows if x[1] == away]
            if row and da.cell(r, 4).value != row[0][6]:
                dmis.append((lbl, da.cell(r, 4).value, row[0][6]))
    check("dashboard_model_spread_matches_engine", not dmis,
          "mismatches: %s" % dmis[:3] if dmis else "16/16 dashboard rows agree with ENGINE")

    # ---------- the five pins ----------
    i_dal, i_nyg = teams.index("DAL"), teams.index("NYG")
    dn = [x for x in rows if x[0] == "2026_01_DAL_NYG"][0]
    for n, ok in (("pin_DAL_-1.07_rank_22",
                   tr.cell(5 + i_dal, 10).value == -1.07 and tr.cell(5 + i_dal, 11).value == 22),
                  ("pin_NYG_-1.90_rank_24",
                   tr.cell(5 + i_nyg, 10).value == -1.9 and tr.cell(5 + i_nyg, 11).value == 24),
                  ("pin_final_margin_+0.77", dn[5] == 0.77),
                  ("pin_fair_line_NYG_-0.8", dn[6] == "NYG -0.8"),
                  ("pin_edge_+3.27_on_NYG_+2.5", dn[8] == 3.27 and dn[9].startswith("NYG"))):
        check(n, ok)

    # ---------- timezone sensitivity (local, no Drive call) ----------
    volatile = []
    for s in nf.sheetnames:
        for row in nf[s].iter_rows():
            for c in row:
                t = ftext(c)
                if t and any(fn in t.upper() for fn in ("NOW(", "TODAY(", "RANDBETWEEN(", "RAND(")):
                    volatile.append((s, c.coordinate))
    check("no_clock_dependent_functions", not volatile,
          "%d found: %s" % (len(volatile), volatile[:5]) if volatile
          else "no NOW/TODAY/RAND anywhere — every date is a static value, so the "
               "spreadsheet time zone cannot change any computed output")

    with open(OUT, "w", newline="") as f:
        w = csv.writer(f, lineterminator="\n")
        w.writerow(["GameID", "Away", "Home", "AwayEff", "HomeEff", "FinalMargin",
                    "Model spread", "MktHomeSpr", "SpreadEdge", "SupportedSide",
                    "ATS_REC", "Reconciliation"])
        w.writerows(rows)

    width = max(len(n) for n, _o, _d in R)
    for n, ok, d in R:
        print("%-4s %-*s %s" % ("PASS" if ok else "FAIL", width, n, d))
    bad_n = [n for n, ok, _d in R if not ok]
    print("\n%d checks, %d failed" % (len(R), len(bad_n)))
    print("reconciliation written to", OUT)
    return 1 if bad_n else 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1]))
