#!/usr/bin/env python3
import openpyxl
WB="TTW_NFL_v1_1_1 Version 2.xlsx"
wb=openpyxl.load_workbook(WB, data_only=False)

def header(sheet, nrows=6, ncols=14):
    ws=wb[sheet]
    print(f"\n===== {sheet} (rows 1-{nrows}) =====")
    for r in range(1,nrows+1):
        cells=[f"{ws.cell(r,c).coordinate}={repr(ws.cell(r,c).value)[:26]}" for c in range(1,ncols+1) if ws.cell(r,c).value is not None]
        if cells: print(f"R{r}:"," | ".join(cells))

for s in ["MARKET LINES","ADJUSTMENTS","QB VALUES","TEAM RATINGS","DATA QUALITY"]:
    header(s, 8, 12)
