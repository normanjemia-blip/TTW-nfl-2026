#!/usr/bin/env python3
"""Diagnose the DATA QUALITY rating-mean #DIV/0! condition. Read-only; proposes a patch, applies nothing."""
import openpyxl, json
A="TTW_NFL_Power_Ratings_2026_v1.1_AUTHORITATIVE.xlsx"
wf=openpyxl.load_workbook(A,data_only=False); wv=openpyxl.load_workbook(A,data_only=True)
ERR={"#DIV/0!","#REF!","#VALUE!","#N/A","#NAME?","#NUM!","#NULL!"}
cells=[]
for sh in wv.worksheets:
    for row in sh.iter_rows():
        for c in row:
            if isinstance(c.value,str) and c.value in ERR:
                cells.append({"sheet":sh.title,"cell":c.coordinate,"cached":c.value,
                              "formula":wf[sh.title][c.coordinate].value})
# Are the source ranges empty? (root-cause test)
calc=wf["CALC"]
ranges={"B39":("F",2,33),"B40":("I",2,33),"B41":("L",2,33),"B42":("O",2,33),"B43":("AC",2,33)}
empt={}
for tgt,(col,r0,r1) in ranges.items():
    vals=[calc[f"{col}{r}"].value for r in range(r0,r1+1)]
    numeric=[v for v in vals if isinstance(v,(int,float))]
    empt[tgt]={"range":f"CALC!{col}{r0}:{col}{r1}","numeric_count":len(numeric),
               "all_formula_or_blank":all(v is None or (isinstance(v,str) and v.startswith("=")) for v in vals)}
# Consumers: does anything gate on DATA QUALITY!B8 / CALC B39:B43?
consumers=[]
for sh in wf.worksheets:
    for row in sh.iter_rows():
        for c in row:
            v=c.value
            if isinstance(v,str) and v.startswith("="):
                if "CALC!$B$42" in v or "CALC!$B$39" in v or "CALC!$B$43" in v:
                    consumers.append(f"{sh.title}!{c.coordinate}: {v[:90]}")
print(json.dumps({"error_cells":cells,"root_cause_ranges":empt,"consumers_of_mean_checks":consumers},indent=2))
