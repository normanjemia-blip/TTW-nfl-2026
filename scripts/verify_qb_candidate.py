#!/usr/bin/env python3
import openpyxl, zipfile, hashlib, json, csv, datetime
V12="TTW_NFL_Power_Ratings_2026_v1.2_QB_WORKING.xlsx"
CAND="TTW_NFL_Power_Ratings_2026_v1.2.1_QB_CANDIDATE.xlsx"
AUTH="TTW_NFL_Power_Ratings_2026_v1.1_AUTHORITATIVE.xlsx"
def sha(p): return hashlib.sha256(open(p,"rb").read()).hexdigest()

rep={"v1_2_sha256":sha(V12),"candidate_sha256":sha(CAND),"authoritative_sha256":sha(AUTH)}

a=openpyxl.load_workbook(V12,data_only=False)
b=openpyxl.load_workbook(CAND,data_only=False)

rep["sheet_count"]=len(b.sheetnames)
rep["sheet_order_identical"]=a.sheetnames==b.sheetnames
rep["visibility_identical"]=[w.sheet_state for w in a.worksheets]==[w.sheet_state for w in b.worksheets]

def fmap(wb):
    m={}
    for sh in wb.worksheets:
        for row in sh.iter_rows():
            for c in row:
                v=c.value
                if (isinstance(v,str) and v.startswith("=")) or c.data_type=="f":
                    m[(sh.title,c.coordinate)]=v.text if hasattr(v,"text") else v
    return m
fa,fb=fmap(a),fmap(b)
rep["formula_count"]=len(fb)
rep["formula_count_is_57399"]=len(fb)==57399
rep["formula_coords_identical"]=set(fa)==set(fb)
rep["formula_text_diffs"]=sum(1 for k in fa if fa[k]!=fb.get(k))

# full cell diff vs v1.2
diffs=[]
for sh in a.sheetnames:
    x=a[sh]; y=b[sh]
    mr=max(x.max_row,y.max_row); mc=max(x.max_column,y.max_column)
    for r in range(1,mr+1):
        for c in range(1,mc+1):
            va=x.cell(r,c).value; vb=y.cell(r,c).value
            va=va.text if hasattr(va,"text") else va
            vb=vb.text if hasattr(vb,"text") else vb
            if va!=vb:
                diffs.append({"sheet":sh,"cell":x.cell(r,c).coordinate,
                              "old":("" if va is None else str(va)),
                              "new":("" if vb is None else str(vb))})
rep["total_cell_diffs_vs_v1_2"]=len(diffs)
rep["diff_sheets"]=sorted(set(d["sheet"] for d in diffs))
# allowed sheets only
rep["only_allowed_sheets_changed"]=set(rep["diff_sheets"])<={"QB VALUES","START HERE","CHANGELOG"}
# prohibited sheets untouched
prohibited=["IMPORT SCHEDULE","MARKET LINES","ADJUSTMENTS","TEAM RATINGS","SETTINGS","HISTORY 2025","BACKTEST",
            "ENGINE","DASHBOARD","DATA QUALITY","IMPORT STATS","MAP","CLEAN","CALC","LISTS","PRESEASON","AUDIT","DICTIONARY"]
rep["prohibited_sheet_changes"]=[d for d in diffs if d["sheet"] in prohibited]

# QB delta + status computation (formula logic; cached values are stale until recalc)
q=b["QB VALUES"]
CURRENT_SEASON=2026
ASOF=datetime.datetime(2026,7,13); STALE=30
rows=[]; nonzero=[]; ok=0; unc=0
for r in range(5,37):
    team=q.cell(r,1).value; C=q.cell(r,3).value; D=q.cell(r,4).value
    E=q.cell(r,5).value; I=q.cell(r,9).value; K=q.cell(r,11).value; M=q.cell(r,13).value
    delta = 0 if (C in (None,"") or E in (None,"")) else round(E-C,4)
    if delta!=0: nonzero.append({"team":team,"delta":delta})
    stale = (K not in (None,"")) and ((ASOF-K).days > STALE if isinstance(K,datetime.datetime) else False)
    flag = 1 if (D in (None,"") or I=="Low" or M!=CURRENT_SEASON or stale) else 0
    status = "UNCERTAIN" if flag==1 else "OK"
    if flag==1: unc+=1
    else: ok+=1
    rows.append({"team":team,"baseline_qb":q.cell(r,2).value,"baseline_value":C,"active_qb":D,
                 "active_value":E,"delta":delta,"confidence":I,"status":status})
rep["qb_nonzero_deltas"]=nonzero
rep["exactly_one_nonzero_delta_LV_0.50"]=(len(nonzero)==1 and nonzero[0]["team"]=="LV" and nonzero[0]["delta"]==0.5)
rep["qb_status_OK"]=ok; rep["qb_status_UNCERTAIN"]=unc
rep["qb_status_matches_29_3"]=(ok==29 and unc==3)
rep["uncertain_teams"]=[x["team"] for x in rows if x["status"]=="UNCERTAIN"]
rep["qb_rows"]=rows

# zip parity
zs=zipfile.ZipFile(V12); zd=zipfile.ZipFile(CAND)
rep["zip_members_changed"]=[n for n in zs.namelist() if zs.read(n)!=zd.read(n)]
rep["drawings_persons_identical"]=all(zs.read(n)==zd.read(n) for n in zs.namelist() if "drawing" in n or "person" in n)

json.dump(rep, open("audit/qb_candidate_verification.json","w"), indent=2, default=str)

# changed cells CSV/JSON
with open("audit/qb_candidate_changed_cells.csv","w",newline="") as f:
    w=csv.DictWriter(f,fieldnames=["sheet","cell","old","new"]); w.writeheader()
    for d in diffs: w.writerow({k:(v[:300] if isinstance(v,str) else v) for k,v in d.items()})
json.dump(diffs, open("audit/qb_candidate_changed_cells.json","w"), indent=2, default=str)

for k in ["candidate_sha256","authoritative_sha256","sheet_count","sheet_order_identical","visibility_identical",
          "formula_count","formula_count_is_57399","formula_coords_identical","formula_text_diffs",
          "total_cell_diffs_vs_v1_2","diff_sheets","only_allowed_sheets_changed","prohibited_sheet_changes",
          "qb_nonzero_deltas","exactly_one_nonzero_delta_LV_0.50","qb_status_OK","qb_status_UNCERTAIN",
          "qb_status_matches_29_3","uncertain_teams","zip_members_changed","drawings_persons_identical"]:
    print(f"{k}: {rep[k]}")
print("\n--- changed cells ---")
for d in diffs: print(f"  [{d['sheet']}!{d['cell']}] {d['old'][:40]!r} -> {d['new'][:55]!r}")
