#!/usr/bin/env python3
import openpyxl
WB="TTW_NFL_v1_1_1 Version 2.xlsx"
wb=openpyxl.load_workbook(WB, data_only=False)
wbv=openpyxl.load_workbook(WB, data_only=True)

# SETTINGS full dump
ws=wb["SETTINGS"]
print("===== SETTINGS =====")
for r in range(1, ws.max_row+1):
    cells=[f"{ws.cell(r,c).coordinate}={repr(ws.cell(r,c).value)[:50]}" for c in range(1,5) if ws.cell(r,c).value is not None]
    if cells: print(" | ".join(cells))

# IMPORT SCHEDULE header + season/game_type breakdown
ws=wb["IMPORT SCHEDULE"]
print("\n===== IMPORT SCHEDULE header (row 1-6) =====")
for r in range(1,7):
    cells=[f"{ws.cell(r,c).coordinate}={repr(ws.cell(r,c).value)[:22]}" for c in range(1,12) if ws.cell(r,c).value is not None]
    if cells: print(f"R{r}:", " | ".join(cells))
