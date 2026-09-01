#!/usr/bin/env python3
import openpyxl, zipfile, hashlib, json, csv, datetime, re
BASE="TTW_NFL_Power_Ratings_2026_v1.2.1_QB_CANDIDATE.xlsx"
CAND="TTW_NFL_Power_Ratings_2026_v1.2.2_QB_SWEEP_CANDIDATE.xlsx"
AUTH="TTW_NFL_Power_Ratings_2026_v1.1_AUTHORITATIVE.xlsx"
def sha(p): return hashlib.sha256(open(p,"rb").read()).hexdigest()
rep={"base_sha256":sha(BASE),"candidate_sha256":sha(CAND),"authoritative_sha256":sha(AUTH)}

a=openpyxl.load_workbook(BASE,data_only=False); b=openpyxl.load_workbook(CAND,data_only=False)
rep["sheet_count"]=len(b.sheetnames); rep["sheet_count_is_21"]=len(b.sheetnames)==21
rep["sheet_order_identical"]=a.sheetnames==b.sheetnames
rep["visibility_identical"]=[w.sheet_state for w in a.worksheets]==[w.sheet_state for w in b.worksheets]

def fmap(wb):
    m={}
    for sh in wb.worksheets:
        for row in sh.iter_rows():
            for c in row:
                v=c.value
                if (isinstance(v,str) and v.startswith("=")) or c.data_type=="f":
                    m[(sh.title,c.coordinate)]=v.text if hasattr(v,"text") else v
    return m
fa,fb=fmap(a),fmap(b)
rep["formula_count"]=len(fb); rep["formula_count_is_57399"]=len(fb)==57399
rep["formula_coords_identical"]=set(fa)==set(fb)
rep["formula_text_diffs"]=sum(1 for k in fa if fa[k]!=fb.get(k))

diffs=[]
for sh in a.sheetnames:
    x=a[sh]; y=b[sh]
    mr=max(x.max_row,y.max_row); mc=max(x.max_column,y.max_column)
    for r in range(1,mr+1):
        for c in range(1,mc+1):
            va=x.cell(r,c).value; vb=y.cell(r,c).value
            va=va.text if hasattr(va,"text") else va
            vb=vb.text if hasattr(vb,"text") else vb
            if va!=vb:
                diffs.append({"sheet":sh,"cell":x.cell(r,c).coordinate,
                              "old":("" if va is None else str(va)),"new":("" if vb is None else str(vb))})
rep["total_cell_diffs"]=len(diffs)
rep["diff_sheets"]=sorted(set(d["sheet"] for d in diffs))
rep["only_allowed_sheets_changed"]=set(rep["diff_sheets"])<={"QB VALUES","START HERE","CHANGELOG"}
for nm,sheet in [("market_lines_changes","MARKET LINES"),("adjustments_changes","ADJUSTMENTS"),
                 ("team_ratings_changes","TEAM RATINGS"),("settings_changes","SETTINGS"),
                 ("schedule_changes","IMPORT SCHEDULE"),("history_changes","HISTORY 2025"),
                 ("backtest_changes","BACKTEST"),("preseason_changes","PRESEASON")]:
    rep[nm]=[d for d in diffs if d["sheet"]==sheet]
# QB diffs confined to authorized rows
AUTH_ROWS={6,12,23,25}
badrows=[d for d in diffs if d["sheet"]=="QB VALUES" and int(re.match(r'[A-Z]+(\d+)$',d["cell"]).group(1)) not in AUTH_ROWS]
rep["qb_diffs_outside_authorized_rows"]=badrows

# statuses
q=b["QB VALUES"]; CUR=2026; ASOF=datetime.datetime(2026,7,13); STALE=30
ok=unc=0; nonzero=[]; uncertain=[]
for r in range(5,37):
    t=q.cell(r,1).value; C=q.cell(r,3).value; D=q.cell(r,4).value
    E=q.cell(r,5).value; I=q.cell(r,9).value; Kv=q.cell(r,11).value; M=q.cell(r,13).value
    delta=0 if (C in (None,"") or E in (None,"")) else round(E-C,4)
    if delta!=0: nonzero.append({"team":t,"baseline":q.cell(r,2).value,"active":D,"delta":delta})
    stale=isinstance(Kv,datetime.datetime) and (ASOF-Kv).days>STALE
    flag=1 if (D in (None,"") or I=="Low" or M!=CUR or stale) else 0
    if flag: unc+=1; uncertain.append(t)
    else: ok+=1
rep["qb_OK"]=ok; rep["qb_UNCERTAIN"]=unc; rep["uncertain_teams"]=uncertain
rep["nonzero_deviations"]=nonzero
rep["exactly_one_nonzero_LV_050"]=(len(nonzero)==1 and nonzero[0]["team"]=="LV" and nonzero[0]["delta"]==0.5)
rep["counts_31_1"]=(ok==31 and unc==1)

zs=zipfile.ZipFile(BASE); zd=zipfile.ZipFile(CAND)
rep["zip_members_changed"]=[n for n in zs.namelist() if zs.read(n)!=zd.read(n)]
rep["drawings_persons_identical"]=all(zs.read(n)==zd.read(n) for n in zs.namelist() if "drawing" in n or "person" in n)

json.dump(rep,open("audit/qb_sweep_verification.json","w"),indent=2,default=str)
with open("audit/qb_sweep_changed_cells.csv","w",newline="") as f:
    w=csv.DictWriter(f,fieldnames=["sheet","cell","old","new"]); w.writeheader()
    for d in diffs: w.writerow({k:(v[:300] if isinstance(v,str) else v) for k,v in d.items()})
json.dump(diffs,open("audit/qb_sweep_changed_cells.json","w"),indent=2,default=str)

for k in ["candidate_sha256","authoritative_sha256","sheet_count_is_21","sheet_order_identical","visibility_identical",
          "formula_count","formula_count_is_57399","formula_coords_identical","formula_text_diffs",
          "total_cell_diffs","diff_sheets","only_allowed_sheets_changed","qb_diffs_outside_authorized_rows",
          "market_lines_changes","adjustments_changes","team_ratings_changes","settings_changes","schedule_changes",
          "history_changes","backtest_changes","preseason_changes","qb_OK","qb_UNCERTAIN","uncertain_teams",
          "nonzero_deviations","exactly_one_nonzero_LV_050","counts_31_1","zip_members_changed","drawings_persons_identical"]:
    print(f"{k}: {rep[k]}")
print("\n--- changed cells ---")
for d in diffs: print(f"  [{d['sheet']}!{d['cell']}] {d['old'][:30]!r} -> {d['new'][:45]!r}")
