#!/usr/bin/env python3
import openpyxl
WB="TTW_NFL_v1_1_1 Version 2.xlsx"
wbv=openpyxl.load_workbook(WB, data_only=True)
ERRS={"#REF!","#DIV/0!","#N/A","#VALUE!","#NAME?","#NULL!","#NUM!","#SPILL!","#CALC!"}
total=0; byerr={}
persheet={}
for ws in wbv.worksheets:
    cnt=0
    for row in ws.iter_rows():
        for c in row:
            v=c.value
            if isinstance(v,str) and v in ERRS:
                cnt+=1; total+=1
                byerr[v]=byerr.get(v,0)+1
    if cnt: persheet[ws.title]=cnt
print("TOTAL error cells (cached):", total)
print("By type:", byerr)
print("By sheet:", persheet)

# DATA QUALITY summary computed
wd=wbv["DATA QUALITY"]
print("\nDATA QUALITY summary (rows 5-12):")
for r in range(5,13):
    a=wd.cell(r,1).value; b=wd.cell(r,2).value
    if a is not None: print(f"  {a!r}: {b!r}")

# Version banner
ws=wbv["START HERE"]
print("\nSTART HERE A1:", ws.cell(1,1).value)
print("START HERE A2:", ws.cell(2,1).value)
