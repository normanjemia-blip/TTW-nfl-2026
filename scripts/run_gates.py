#!/usr/bin/env python3
"""Workbook invariant gate suite for the v1.4 authoritative workbook (read-only). Exit 0 = pass.

Re-pinned at the v1.4 release (2026-09-01). Every threshold below is the value production
actually carries after the v1.4 promotion, so a regression to the v1.1 shape — blank
Source B, the old coercing D-column formula, the pre-promotion as-of date, or altered
weights/thresholds — fails explicitly rather than passing quietly.

Drift is measured by semantic content, never by archive bytes: a Google Sheets export is
repackaged on every download, so ZIP SHA equality is meaningless as a live-drift test.
See scripts/semantic_fingerprint.py.
"""
import sys

import openpyxl

A = "TTW_NFL_Power_Ratings_2026_v1.4_AUTHORITATIVE.xlsx"

D_FIXED = ('IFERROR(IF(ISNUMBER(INDEX(CalcGoverned,MATCH($A{r},CalcTeams,0))),'
           'ROUND(INDEX(CalcGoverned,MATCH($A{r},CalcTeams,0)),2),""),"")')
D_OLD = 'IFERROR(ROUND(INDEX(CalcGoverned,MATCH($A{r},CalcTeams,0)),2),"")'
PINS = {"DAL": (-1.07, 22), "NYG": (-1.9, 24)}


def ftext(c):
    v = c.value
    if hasattr(v, "text"):
        return v.text.lstrip("=")
    return v[1:] if isinstance(v, str) and v.startswith("=") else None


def main():
    f = []
    wf = openpyxl.load_workbook(A, data_only=False)
    wv = openpyxl.load_workbook(A, data_only=True)

    # ---- structure ----
    if len(wf.sheetnames) != 22:
        f.append("sheets=%d != 22" % len(wf.sheetnames))
    n = sum(1 for sh in wf.worksheets for row in sh.iter_rows() for c in row
            if ftext(c) is not None)
    if n != 57399:
        f.append("formulas=%d != 57399" % n)
    if wv["START HERE"]["A1"].value != "TO THE WINDOW — NFL POWER RATINGS 2026 (v1.4)":
        f.append("START HERE banner is not v1.4: %r" % wv["START HERE"]["A1"].value)

    # ---- settings: the v1.4 baseline ----
    s = wv["SETTINGS"]
    asof = s.cell(7, 2).value
    asof = asof.strftime("%Y-%m-%d") if hasattr(asof, "strftime") else str(asof)
    if asof != "2026-09-01":
        f.append("as-of date is %s, expected 2026-09-01" % asof)
    if s.cell(5, 2).value != 2026 or s.cell(6, 2).value != 1:
        f.append("season/week drifted: %r/%r" % (s.cell(5, 2).value, s.cell(6, 2).value))
    if s.cell(67, 2).value != "Y":
        f.append("BET labels must be Y at v1.4 (SETTINGS B67=%r)" % s.cell(67, 2).value)
    if [s.cell(26, 2).value, s.cell(27, 2).value, s.cell(28, 2).value] != [1.5, 1.5, 1.0]:
        f.append("ATS thresholds drifted from 1.5/1.5/1.0")
    if [s.cell(29, 2).value, s.cell(30, 2).value, s.cell(31, 2).value] != [3.0, 1.5, 1.0]:
        f.append("Totals thresholds drifted from 3.0/1.5/1.0")
    if s.cell(40, 2).value != "VALIDATE-ONLY":
        f.append("win-totals mode must remain VALIDATE-ONLY")
    if s.cell(9, 2).value != 1.6 or s.cell(38, 2).value != 0.33 or s.cell(39, 2).value != 2.1:
        f.append("HFA / prior regression / pts-per-win drifted")
    if s.cell(44, 2).value != 0.8:
        f.append("week-1 preseason blend weight != 0.80")

    # ---- schedule ----
    sch = wf["IMPORT SCHEDULE"]
    reg = sum(1 for r in range(6, sch.max_row + 1)
              if sch.cell(r, 2).value == 2026 and sch.cell(r, 3).value == "REG")
    if reg != 272:
        f.append("2026 REG games=%d != 272" % reg)

    # ---- PRESEASON: Source B populated, Source C blank, weights unchanged ----
    p, pv = wf["PRESEASON"], wv["PRESEASON"]
    srcb = sum(1 for r in range(5, 37) if pv.cell(r, 9).value not in (None, ""))
    if srcb != 32:
        f.append("PRESEASON Source B populated %d/32, expected 32/32" % srcb)
    if sum(1 for r in range(5, 37) if pv.cell(r, 12).value != "2026-09-01"):
        f.append("PRESEASON Source B as-of must be 2026-09-01 in all 32 rows")
    if sum(1 for r in range(5, 37) if pv.cell(r, 20).value != 2):
        f.append("PRESEASON 'Sources used' must be 2 in all 32 rows")
    srcc = sum(1 for r in range(5, 37) if pv.cell(r, 14).value not in (None, ""))
    if srcc:
        f.append("PRESEASON Source C must remain blank (%d populated)" % srcc)
    badw = [pv.cell(r, 2).value for r in range(5, 37)
            if (pv.cell(r, 8).value, pv.cell(r, 13).value, pv.cell(r, 18).value)
            != (0.4, 0.35, 0.25)]
    if badw:
        f.append("source weights drifted from 0.40/0.35/0.25: %s" % badw)

    # ---- TEAM RATINGS: the corrected D formula and the GP=0 fallback ----
    tf, tv = wf["TEAM RATINGS"], wv["TEAM RATINGS"]
    wrong = [r for r in range(5, 37) if ftext(tf.cell(r, 4)) != D_FIXED.format(r=r)]
    if wrong:
        old = [r for r in wrong if ftext(tf.cell(r, 4)) == D_OLD.format(r=r)]
        f.append("TEAM RATINGS D%s not the ISNUMBER-protected formula%s"
                 % (wrong if len(wrong) < 6 else "5:D36",
                    " (rows %s still carry the old coercing form)" % old if old else ""))
    if sum(1 for r in range(5, 37) if tv.cell(r, 2).value != 0):
        f.append("GP must be 0 for all 32 teams at week 1")
    notblank = [tv.cell(r, 1).value for r in range(5, 37)
                if tv.cell(r, 4).value not in (None, "")]
    if notblank:
        f.append("GP=0 fallback broken: D must be blank, not 0, for %s" % notblank)
    bad_fb = [tv.cell(r, 1).value for r in range(5, 37)
              if not (tv.cell(r, 6).value == tv.cell(r, 8).value == tv.cell(r, 10).value
                      == pv.cell(r, 19).value)]
    if bad_fb:
        f.append("GP=0 fallback broken: F=H=J=prior fails for %s" % bad_fb)

    # ---- the five DAL/NYG pins ----
    idx = {tv.cell(r, 1).value: r for r in range(5, 37)}
    for team, (rating, rank) in PINS.items():
        r = idx[team]
        if (tv.cell(r, 10).value, tv.cell(r, 11).value) != (rating, rank):
            f.append("%s pin failed: %s/%s expected %s/%s"
                     % (team, tv.cell(r, 10).value, tv.cell(r, 11).value, rating, rank))
    en = wv["ENGINE"]
    dn = [r for r in range(5, 300) if en.cell(r, 2).value == "2026_01_DAL_NYG"]
    if not dn:
        f.append("DAL@NYG row not found in ENGINE")
    else:
        r = dn[0]
        got = (round(en.cell(r, 19).value, 2), en.cell(r, 20).value,
               round(en.cell(r, 22).value, 2))
        if got != (0.77, "NYG -0.8", 3.27):
            f.append("DAL@NYG pins failed: FinalMargin/fair line/edge = %r" % (got,))

    if f:
        print("GATE SUITE: FAIL")
        for x in f:
            print("  -", x)
        return 1
    print("GATE SUITE: PASS (v1.4 baseline — 22 sheets, %d formulas, as-of 2026-09-01, "
          "SrcB 32/32, SrcC blank, weights .40/.35/.25, ATS 1.5/1.5/1.0, BET labels Y, "
          "VALIDATE-ONLY, 272 REG, D5:D36 ISNUMBER-protected, GP=0 fallback 32/32, "
          "DAL -1.07/r22, NYG -1.90/r24, DAL@NYG +0.77 / NYG -0.8 / +3.27)" % n)
    return 0


if __name__ == "__main__":
    sys.exit(main())
