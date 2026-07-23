#!/usr/bin/env python3
import openpyxl
WB="TTW_NFL_v1_1_1 Version 2.xlsx"
wbf=openpyxl.load_workbook(WB, data_only=False)
wbv=openpyxl.load_workbook(WB, data_only=True)
for sheet in ["CALC","DATA QUALITY"]:
    wf=wbf[sheet]; wv=wbv[sheet]
    print(f"\n===== {sheet}: #DIV/0! cells =====")
    for row in wv.iter_rows():
        for c in row:
            if c.value=="#DIV/0!":
                coord=c.coordinate
                print(f"{coord}: formula = {wf[coord].value}")
# Show CALC context rows 40-44 labels
wf=wbf["CALC"]; wv=wbv["CALC"]
print("\n--- CALC rows 40-44 (labels col A + values) ---")
for r in range(38,45):
    cells=[f"{wf.cell(r,c).coordinate}={repr(wf.cell(r,c).value)[:45]}" for c in range(1,6) if wf.cell(r,c).value is not None]
    if cells: print(f"R{r}:"," | ".join(cells))
