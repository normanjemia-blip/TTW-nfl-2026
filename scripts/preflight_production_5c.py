#!/usr/bin/env python3
"""Phase 5C production preflight — READ-ONLY.

Compares a fresh production export against the committed promotion-manifest baseline
and confirms all 133 directly-entered target cells still hold their expected
pre-promotion values.

Usage: preflight_production_5c.py <fresh_production_export.xlsx>

Writes nothing to Google. Exits non-zero on any drift.
"""
import csv
import os
import sys

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
BASE = os.path.join(ROOT, "TTW_NFL_2026_LIVE_EXPORT_20260901_BASE.xlsx")
MANIFEST = os.path.join(ROOT, "audit", "TTW_NFL_2026_Promotion_Manifest_v14_20260901.csv")

R = []
def check(name, ok, detail=""):
    R.append((name, bool(ok), detail))


def ftext(c):
    v = c.value
    if hasattr(v, "text"):
        return v.text.lstrip("=")
    return v[1:] if isinstance(v, str) and v.startswith("=") else None


def maps(path):
    wb = openpyxl.load_workbook(path)
    f, k = {}, {}
    for s in wb.sheetnames:
        for row in wb[s].iter_rows():
            for c in row:
                t = ftext(c)
                if t is not None:
                    f[(s, c.coordinate)] = t
                elif c.value is not None:
                    k[(s, c.coordinate)] = c.value
    wb.close()
    return wb.sheetnames if False else (f, k)


def main(fresh):
    fb, cb = maps(BASE)
    ff, cf = maps(fresh)
    wf = openpyxl.load_workbook(fresh)
    wv = openpyxl.load_workbook(fresh, data_only=True)
    wb_base = openpyxl.load_workbook(BASE, data_only=True)

    # ---------- identity ----------
    check("tab_count_and_order_match_baseline",
          wf.sheetnames == openpyxl.load_workbook(BASE).sheetnames,
          "%d tabs" % len(wf.sheetnames))

    # ---------- no drift since the manifest baseline ----------
    check("formula_count_matches_baseline", len(ff) == len(fb) == 57399,
          "fresh %d / baseline %d" % (len(ff), len(fb)))
    fdiff = [k for k in set(fb) | set(ff) if fb.get(k) != ff.get(k)]
    check("no_formula_drift", not fdiff,
          "%d differing formulas: %s" % (len(fdiff), fdiff[:5]) if fdiff else "0 of 57,399")
    kdiff = [k for k in set(cb) | set(cf) if cb.get(k) != cf.get(k)]
    check("no_constant_drift", not kdiff,
          "%d differing constants: %s" % (len(kdiff), kdiff[:5]) if kdiff else "0")

    # ---------- the 133 target cells still hold pre-promotion values ----------
    rows = [r for r in csv.DictReader(open(MANIFEST))
            if r["Action"] in ("PASTE VALUE", "PASTE TEXT", "FORMULA", "TEXT")]
    check("manifest_lists_133_entered_cells", len(rows) == 133, str(len(rows)))
    bad = []
    for r in rows:
        sh, cell, action, before = r["Sheet"], r["Cell / Range"], r["Action"], r["Before"]
        if action == "FORMULA":
            got = ftext(wf[sh][cell])
            want = before.lstrip("=")
            if got != want:
                bad.append("%s!%s formula is %r" % (sh, cell, got))
        else:
            v = wv[sh][cell].value
            got = "(blank)" if v is None else str(v)
            if hasattr(v, "strftime"):
                got = v.strftime("%Y-%m-%d")
            if got != before:
                bad.append("%s!%s is %r, expected %r" % (sh, cell, got, before))
    check("all_133_target_cells_at_expected_pre_promotion_values", not bad,
          "%d unexpected: %s" % (len(bad), bad[:5]) if bad else "133/133 match the manifest 'Before' column")

    # ---------- guardrails still as the manifest assumes ----------
    ps = wv["PRESEASON"]
    check("srcB_still_blank_all_32",
          all(ps.cell(r, 9).value is None for r in range(5, 37)), "ready to receive Source B")
    check("srcC_still_blank_all_32", all(ps.cell(r, 14).value is None for r in range(5, 37)))
    tr = wv["TEAM RATINGS"]
    check("gp_zero_all_32", all(tr.cell(r, 2).value == 0 for r in range(5, 37)))
    check("D_column_still_coerces_to_zero",
          all(tr.cell(r, 4).value == 0 for r in range(5, 37)),
          "32/32 read 0 — the defect is still present, as the manifest assumes")
    st = {str(wv["SETTINGS"].cell(r, 1).value).strip(): wv["SETTINGS"].cell(r, 2).value
          for r in range(1, 80) if wv["SETTINGS"].cell(r, 1).value is not None}
    check("win_totals_mode_validate_only",
          st.get("Win-totals mode (VALIDATE-ONLY until conversion is verified)") == "VALIDATE-ONLY")
    check("banner_still_v11",
          wv["START HERE"]["A1"].value == "TO THE WINDOW — NFL POWER RATINGS 2026 (v1.1)",
          repr(wv["START HERE"]["A1"].value))
    ch = wv["CHANGELOG"]
    check("changelog_row7_still_empty",
          all(ch.cell(7, c).value is None for c in range(1, 5)),
          "ready to receive the v1.4 entry")

    # ---------- the values the promotion will replace ----------
    dal = [i for i in range(32) if tr.cell(5 + i, 1).value == "DAL"][0]
    nyg = [i for i in range(32) if tr.cell(5 + i, 1).value == "NYG"][0]
    check("preflight_DAL_at_-2.26_rank_27",
          tr.cell(5 + dal, 10).value == -2.26 and tr.cell(5 + dal, 11).value == 27)
    check("preflight_NYG_at_-1.59_rank_24",
          tr.cell(5 + nyg, 10).value == -1.59 and tr.cell(5 + nyg, 11).value == 24)

    width = max(len(n) for n, _o, _d in R)
    for n, ok, d in R:
        print("%-4s %-*s %s" % ("PASS" if ok else "FAIL", width, n, d))
    bad_n = [n for n, ok, _d in R if not ok]
    print("\n%d checks, %d failed" % (len(R), len(bad_n)))
    return 1 if bad_n else 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1]))
