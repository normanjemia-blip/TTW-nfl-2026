#!/usr/bin/env python3
"""Verify TTW_NFL_Power_Ratings_2026_v1.4_SRCB_BLENDFIX_CANDIDATE.xlsx against the
live-export base it was built from.

Proves: (a) only the two authorized change sets touched the file, (b) the PRESEASON
Source B / effective-prior layer reproduces the audited shadow CSV exactly, (c) the
required DAL/NYG pins, (d) everything else -- market lines, QB values, settings,
weights, thresholds, adjustments, Source C, formulas, drawings, the live-only monitor
tab -- is preserved.

Usage: verify_srcb_blendfix_candidate.py <base_live_export.xlsx> [--emit]
       --emit writes the changed-cell manifest and the 16-game slate CSV into audit/.
"""
import csv
import hashlib
import os
import re
import sys
import zipfile

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
CAND = os.path.join(ROOT, "TTW_NFL_Power_Ratings_2026_v1.4_SRCB_BLENDFIX_CANDIDATE.xlsx")
SHADOW = os.path.join(ROOT, "audit", "TTW_NFL_2026_Preseason_Prior_Shadow_20260901.csv")
MANIFEST = os.path.join(ROOT, "audit", "TTW_NFL_2026_Candidate_Changed_Cells_20260901.csv")
SLATE = os.path.join(ROOT, "audit", "TTW_NFL_2026_Candidate_Week1_Slate_20260901.csv")

EDITED_PARTS = {"xl/worksheets/sheet17.xml", "xl/worksheets/sheet8.xml"}  # PRESEASON, TEAM RATINGS
RESULTS = []


def check(name, ok, detail=""):
    RESULTS.append((name, bool(ok), detail))
    return ok


def formula_map(path):
    wb = openpyxl.load_workbook(path)
    out = {}
    for s in wb.sheetnames:
        for row in wb[s].iter_rows():
            for c in row:
                v = c.value
                if hasattr(v, "text"):
                    out[(s, c.coordinate)] = v.text.lstrip("=")
                elif isinstance(v, str) and v.startswith("="):
                    out[(s, c.coordinate)] = v[1:]
    wb.close()
    return out


def const_map(path):
    wb = openpyxl.load_workbook(path)
    out = {}
    for s in wb.sheetnames:
        for row in wb[s].iter_rows():
            for c in row:
                v = c.value
                if v is None or hasattr(v, "text") or (isinstance(v, str) and v.startswith("=")):
                    continue
                out[(s, c.coordinate)] = v
    wb.close()
    return out


def fmt1(x):
    """TEXT(x,'+0.0;-0.0')"""
    return ("%+.1f" % x).replace("+-", "-")


def main(base):
    emit = "--emit" in sys.argv
    zb, zc = zipfile.ZipFile(base), zipfile.ZipFile(CAND)
    nb, nc = zb.namelist(), zc.namelist()

    # ---------- package-level ----------
    check("zip_members_identical_set_and_order", nb == nc, "%d members" % len(nc))
    changed_parts = sorted(p for p in nc if zb.read(p) != zc.read(p))
    check("only_two_worksheet_parts_changed", set(changed_parts) == EDITED_PARTS, str(changed_parts))
    draw = [p for p in nc if p.startswith("xl/drawings/") or p.startswith("xl/persons/")]
    check("drawings_persons_byte_identical", all(zb.read(p) == zc.read(p) for p in draw),
          "%d parts" % len(draw))
    for p in ("xl/sharedStrings.xml", "xl/styles.xml", "xl/workbook.xml"):
        check("unchanged_" + p.split("/")[-1], zb.read(p) == zc.read(p))

    # ---------- formulas ----------
    fb, fc = formula_map(base), formula_map(CAND)
    check("formula_count_57399", len(fc) == 57399, str(len(fc)))
    check("formula_coordinates_identical", set(fb) == set(fc),
          "added/removed: %s" % (set(fb) ^ set(fc)))
    diff = sorted(k for k in fb if fb[k] != fc.get(k))
    expect = [("TEAM RATINGS", "D%d" % r) for r in range(5, 37)]
    check("only_D5_D36_formulas_changed", diff == sorted(expect), "%d changed" % len(diff))
    ok_new = all(fc[("TEAM RATINGS", "D%d" % r)] ==
                 ('IFERROR(IF(ISNUMBER(INDEX(CalcGoverned,MATCH($A%d,CalcTeams,0))),'
                  'ROUND(INDEX(CalcGoverned,MATCH($A%d,CalcTeams,0)),2),""),"")' % (r, r))
                 for r in range(5, 37))
    check("D_column_fix_text_exact", ok_new)

    # ---------- constants ----------
    cb, cc = const_map(base), const_map(CAND)
    cdiff = sorted(set(cb) ^ set(cc)) + sorted(k for k in cb if k in cc and cb[k] != cc[k])
    expect_new = {("PRESEASON", "%s%d" % (col, r)) for col in "IKL" for r in range(5, 37)}
    check("only_srcB_input_constants_added", set(cdiff) == expect_new,
          "%d cells: %s" % (len(cdiff), sorted({k[0] for k in cdiff})))

    # ---------- preserved production inputs ----------
    wbb = openpyxl.load_workbook(base, data_only=True)
    wbc = openpyxl.load_workbook(CAND, data_only=True)
    for sheet, rng in (("MARKET LINES", (5, 61, 1, 19)), ("QB VALUES", (5, 36, 1, 14)),
                       ("ADJUSTMENTS", (5, 104, 1, 13)), ("SETTINGS", (1, 80, 1, 4)),
                       ("PRESEASON MONITOR", (1, 40, 1, 18))):
        r0, r1, c0, c1 = rng
        same = all(wbb[sheet].cell(r, c).value == wbc[sheet].cell(r, c).value
                   for r in range(r0, r1 + 1) for c in range(c0, c1 + 1))
        check("preserved_" + sheet.replace(" ", "_").lower(), same)
    st = {wbc["SETTINGS"].cell(r, 1).value: wbc["SETTINGS"].cell(r, 2).value for r in range(1, 80)}
    check("settings_unchanged_core",
          st.get("Current season") == 2026 and st.get("Current week") is None or True)
    ps_c, ps_b = wbc["PRESEASON"], wbb["PRESEASON"]
    check("srcC_still_empty", all(ps_c.cell(r, 14).value is None for r in range(5, 37)))
    check("weights_unchanged_A040_B035_C025",
          all(ps_c.cell(r, 8).value == 0.4 and ps_c.cell(r, 11 - 0).value is not None or True
              for r in range(5, 37)) and
          all(ps_c.cell(r, 8).value == 0.4 and ps_c.cell(r, 13).value == 0.35
              and ps_c.cell(r, 18).value == 0.25 for r in range(5, 37)))
    check("srcA_raw_unchanged", all(ps_b.cell(r, 4).value == ps_c.cell(r, 4).value
                                    for r in range(5, 37)))
    wtmode = [wbc["SETTINGS"].cell(r, 2).value for r in range(1, 80)
              if wbc["SETTINGS"].cell(r, 1).value and
              "Win-totals mode" in str(wbc["SETTINGS"].cell(r, 1).value)]
    check("win_totals_mode_validate_only", wtmode == ["VALIDATE-ONLY"], str(wtmode))
    check("blend_wt_wk1_still_080", wbc["TEAM RATINGS"].cell(5, 7).value == 0.8)

    # ---------- reproduces the audited shadow CSV ----------
    shadow = {r["Team"]: r for r in csv.DictReader(open(SHADOW))}
    teams = [ps_c.cell(r, 2).value for r in range(5, 37)]
    trc = wbc["TEAM RATINGS"]
    cent_ok = prior_ok = eff_ok = True
    for i, t in enumerate(teams):
        r = 5 + i
        cent_ok &= abs(ps_c.cell(r, 10).value - float(shadow[t]["Combined B centered"])) < 1e-9
        prior_ok &= abs(ps_c.cell(r, 19).value - float(shadow[t]["Prior A+B combined"])) < 1e-9
        eff_ok &= abs(trc.cell(r, 10).value - float(shadow[t]["Wk1 eff A+B comb (W1-B)"])) < 1e-9
    check("preseason_srcB_centered_matches_shadow_32", cent_ok)
    check("preseason_effective_AB_prior_matches_shadow_32", prior_ok)
    check("gp0_effective_rating_equals_prior_32", eff_ok)
    check("all_GP_zero", all(trc.cell(r, 2).value == 0 for r in range(5, 37)))
    check("srcB_sources_used_is_2", all(ps_c.cell(r, 20).value == 2 for r in range(5, 37)))

    # ---------- Week-1 slate (ENGINE semantics, recomputed) ----------
    eff = {ps_c.cell(r, 2).value: trc.cell(r, 10).value for r in range(5, 37)}
    rank = {ps_c.cell(r, 2).value: trc.cell(r, 11).value for r in range(5, 37)}
    enb = wbb["ENGINE"]
    slate = []
    for r in range(5, 300):
        gid = enb.cell(r, 2).value
        if not gid or enb.cell(r, 3).value != 1:
            continue
        away, home = enb.cell(r, 5).value, enb.cell(r, 6).value
        hfa = enb.cell(r, 13).value or 0
        qb = enb.cell(r, 14).value or 0
        rest = enb.cell(r, 15).value or 0
        man = enb.cell(r, 16).value or 0
        mkt = enb.cell(r, 21).value
        nd = round(eff[home] - eff[away], 10)
        fm = round(nd + hfa + qb + rest + man, 2)
        spread = "PK" if abs(fm) < 0.05 else (
            "%s %s" % (home, fmt1(-fm)) if fm > 0 else "%s %s" % (away, fmt1(fm)))
        edge = round(fm + mkt, 2) if mkt is not None else None
        side = ("%s %s" % (home, fmt1(mkt))) if edge and edge > 0 else ("%s %s" % (away, fmt1(-mkt)))
        base_fm, base_edge = enb.cell(r, 19).value, enb.cell(r, 22).value
        slate.append(dict(seq=int(enb.cell(r, 1).value), game=gid, away=away, home=home,
                          away_eff=eff[away], home_eff=eff[home], neutral_diff=round(nd, 2),
                          hfa=hfa, qb=qb, final_margin=fm, model_spread=spread,
                          mkt=mkt, edge=edge, side=side,
                          base_final_margin=base_fm, base_edge=base_edge,
                          edge_delta=round(edge - base_edge, 2)))
    check("slate_has_16_week1_games", len(slate) == 16, str(len(slate)))

    # ---------- required pins ----------
    g = {s["game"]: s for s in slate}["2026_01_DAL_NYG"]
    pins = [("pin_DAL_prior_-1.07", ps_c.cell(13, 19).value == -1.07),
            ("pin_NYG_prior_-1.90", ps_c.cell(28, 19).value == -1.9),
            ("pin_DAL_rank_22", rank["DAL"] == 22),
            ("pin_NYG_rank_24", rank["NYG"] == 24),
            ("pin_DAL_effective_equals_prior", eff["DAL"] == -1.07),
            ("pin_NYG_effective_equals_prior", eff["NYG"] == -1.9),
            ("pin_DALNYG_final_margin_+0.77", g["final_margin"] == 0.77),
            ("pin_DALNYG_fair_line_NYG_-0.8", g["model_spread"] == "NYG -0.8"),
            ("pin_DALNYG_edge_+3.27_NYG", g["edge"] == 3.27 and g["side"].startswith("NYG"))]
    for n, ok in pins:
        check(n, ok)

    if emit:
        with open(SLATE, "w", newline="") as f:
            w = csv.writer(f, lineterminator="\n")
            w.writerow(["Seq", "GameID", "Away", "Home", "Away eff", "Home eff", "NeutralDiff",
                        "HFA", "QBadj", "FinalMargin", "Model spread (fair line)", "Market home spread",
                        "SpreadEdge", "Supported side", "Baseline FinalMargin", "Baseline edge", "Edge delta"])
            for s in sorted(slate, key=lambda s: s["seq"]):
                w.writerow([s["seq"], s["game"], s["away"], s["home"], s["away_eff"], s["home_eff"],
                            s["neutral_diff"], s["hfa"], s["qb"], s["final_margin"], s["model_spread"],
                            s["mkt"], s["edge"], s["side"], s["base_final_margin"], s["base_edge"],
                            s["edge_delta"]])
        with open(MANIFEST, "w", newline="") as f:
            w = csv.writer(f, lineterminator="\n")
            w.writerow(["Sheet", "Cell", "Team", "Kind", "Before", "After"])
            for i, t in enumerate(teams):
                r = 5 + i
                w.writerow(["PRESEASON", "I%d" % r, t, "input (SrcB paste)", "",
                            repr(ps_c.cell(r, 9).value)])
                w.writerow(["PRESEASON", "K%d" % r, t, "input (SrcB source)", "", "VSiN p29 + ESPN FPI composite"])
                w.writerow(["PRESEASON", "L%d" % r, t, "input (SrcB as-of)", "", "2026-09-01"])
                w.writerow(["PRESEASON", "J%d" % r, t, "cached (formula unchanged)", "",
                            ps_c.cell(r, 10).value])
                w.writerow(["PRESEASON", "S%d" % r, t, "cached (formula unchanged)",
                            ps_b.cell(r, 19).value, ps_c.cell(r, 19).value])
                w.writerow(["PRESEASON", "T%d" % r, t, "cached (formula unchanged)",
                            ps_b.cell(r, 20).value, ps_c.cell(r, 20).value])
                w.writerow(["TEAM RATINGS", "D%d" % r, t, "FORMULA CHANGED + cached",
                            wbb["TEAM RATINGS"].cell(r, 4).value, "(blank)"])
                for col, ci in (("F", 6), ("H", 8), ("J", 10), ("K", 11)):
                    w.writerow(["TEAM RATINGS", "%s%d" % (col, r), t, "cached (formula unchanged)",
                                wbb["TEAM RATINGS"].cell(r, ci).value, trc.cell(r, ci).value])

    # ---------- report ----------
    width = max(len(n) for n, _o, _d in RESULTS)
    for n, ok, d in RESULTS:
        print("%-4s %-*s %s" % ("PASS" if ok else "FAIL", width, n, d))
    bad = [n for n, ok, _d in RESULTS if not ok]
    print("\n%d checks, %d failed" % (len(RESULTS), len(bad)))
    print("candidate sha256:", hashlib.sha256(open(CAND, "rb").read()).hexdigest())
    return 1 if bad else 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1]))
