#!/usr/bin/env python3
"""Semantic fingerprint of a workbook — the drift test that survives Google's repackaging.

A Google Sheets export is repackaged on every download, so its ZIP SHA-256 changes even
when nothing in the Sheet did. Byte equality is therefore useless as a live-drift test.
This module fingerprints what actually matters: sheet names, every formula by coordinate
and text, and every non-formula constant by coordinate and value.

Two workbooks with the same fingerprint are the same model, whatever the archive bytes say.

CLI: semantic_fingerprint.py <a.xlsx> [b.xlsx]
     one argument  -> print the fingerprint digests
     two arguments -> diff them and exit non-zero on any difference
"""
import hashlib
import json
import sys

import openpyxl


def _ftext(c):
    v = c.value
    if hasattr(v, "text"):
        return v.text.lstrip("=")
    return v[1:] if isinstance(v, str) and v.startswith("=") else None


def fingerprint(path):
    """Return {sheets, formulas, constants, digests} for a workbook."""
    wf = openpyxl.load_workbook(path)
    wv = openpyxl.load_workbook(path, data_only=True)
    formulas, constants = {}, {}
    for s in wf.sheetnames:
        for row in wf[s].iter_rows():
            for c in row:
                t = _ftext(c)
                if t is not None:
                    formulas["%s!%s" % (s, c.coordinate)] = t
                else:
                    v = wv[s][c.coordinate].value
                    if v is not None:
                        constants["%s!%s" % (s, c.coordinate)] = (
                            v.strftime("%Y-%m-%d") if hasattr(v, "strftime") else
                            (round(v, 10) if isinstance(v, float) else v))
    sheets = list(wf.sheetnames)
    wf.close()
    wv.close()

    def dig(obj):
        return hashlib.sha256(
            json.dumps(obj, sort_keys=True, default=str).encode()).hexdigest()

    return {"sheets": sheets, "formulas": formulas, "constants": constants,
            "digests": {"sheets": dig(sheets), "formulas": dig(formulas),
                        "constants": dig(constants),
                        "combined": dig([dig(sheets), dig(formulas), dig(constants)])}}


def diff(a, b, label_a="A", label_b="B"):
    """Return a list of human-readable differences."""
    out = []
    if a["sheets"] != b["sheets"]:
        out.append("sheet names/order differ: %s vs %s" % (a["sheets"], b["sheets"]))
    for kind in ("formulas", "constants"):
        ka, kb = a[kind], b[kind]
        only_a = sorted(set(ka) - set(kb))
        only_b = sorted(set(kb) - set(ka))
        changed = sorted(k for k in set(ka) & set(kb) if ka[k] != kb[k])
        for k in only_a[:20]:
            out.append("%s only in %s: %s = %r" % (kind[:-1], label_a, k, ka[k]))
        for k in only_b[:20]:
            out.append("%s only in %s: %s = %r" % (kind[:-1], label_b, k, kb[k]))
        for k in changed[:20]:
            out.append("%s differs at %s: %s=%r %s=%r" % (kind[:-1], k, label_a, ka[k],
                                                          label_b, kb[k]))
        extra = len(only_a) + len(only_b) + len(changed) - min(20, len(only_a)) \
            - min(20, len(only_b)) - min(20, len(changed))
        if extra > 0:
            out.append("... and %d more %s differences" % (extra, kind))
    return out


if __name__ == "__main__":
    fa = fingerprint(sys.argv[1])
    if len(sys.argv) == 2:
        print("sheets    :", len(fa["sheets"]))
        print("formulas  :", len(fa["formulas"]))
        print("constants :", len(fa["constants"]))
        for k, v in fa["digests"].items():
            print("%-10s: %s" % (k, v))
        sys.exit(0)
    fb = fingerprint(sys.argv[2])
    d = diff(fa, fb, sys.argv[1], sys.argv[2])
    if not d:
        print("SEMANTICALLY IDENTICAL")
        print("  sheets %d | formulas %d | constants %d"
              % (len(fa["sheets"]), len(fa["formulas"]), len(fa["constants"])))
        print("  combined digest:", fa["digests"]["combined"])
        sys.exit(0)
    print("SEMANTIC DIFFERENCES (%d):" % len(d))
    for line in d:
        print(" ", line)
    sys.exit(1)
