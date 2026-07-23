#!/usr/bin/env python3
import openpyxl
from collections import Counter, defaultdict
WB="TTW_NFL_v1_1_1 Version 2.xlsx"
wb=openpyxl.load_workbook(WB, data_only=False)

ws=wb["IMPORT SCHEDULE"]
# columns: A game_id, B season, C game_type, D week, I away_score, K home_score
by_season=Counter()
by_season_type=Counter()
scored=defaultdict(int)   # rows with both scores present
unscored=defaultdict(int)
rows=0
for r in range(6, ws.max_row+1):
    gid=ws.cell(r,1).value
    season=ws.cell(r,2).value
    gtype=ws.cell(r,3).value
    if gid is None or season is None: 
        continue
    rows+=1
    by_season[season]+=1
    by_season_type[(season,gtype)]+=1
    asc=ws.cell(r,9).value
    hsc=ws.cell(r,11).value
    if asc is not None and hsc is not None:
        scored[season]+=1
    else:
        unscored[season]+=1

print("Total schedule data rows:", rows)
print("\nBy season:", dict(by_season))
print("\nBy (season, game_type):")
for k in sorted(by_season_type):
    print("  ", k, by_season_type[k])
print("\nScored (both scores present) by season:", dict(scored))
print("Unscored by season:", dict(unscored))

# PRESEASON sheet
ws=wb["PRESEASON"]
print("\n===== PRESEASON (first 20 rows) =====")
for r in range(1, 20):
    cells=[f"{ws.cell(r,c).coordinate}={repr(ws.cell(r,c).value)[:30]}" for c in range(1,10) if ws.cell(r,c).value is not None]
    if cells: print(" | ".join(cells))

# HISTORY 2025
ws=wb["HISTORY 2025"]
print("\n===== HISTORY 2025 (first 15 rows) =====")
for r in range(1, 15):
    cells=[f"{ws.cell(r,c).coordinate}={repr(ws.cell(r,c).value)[:30]}" for c in range(1,10) if ws.cell(r,c).value is not None]
    if cells: print(" | ".join(cells))

# BACKTEST
ws=wb["BACKTEST"]
print("\n===== BACKTEST (first 25 rows) =====")
for r in range(1, 25):
    cells=[f"{ws.cell(r,c).coordinate}={repr(ws.cell(r,c).value)[:35]}" for c in range(1,8) if ws.cell(r,c).value is not None]
    if cells: print(" | ".join(cells))
