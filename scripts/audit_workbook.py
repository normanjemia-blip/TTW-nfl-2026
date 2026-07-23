#!/usr/bin/env python3
"""NFL Preseason Readiness — Baseline Audit grounding script.
Reads the workbook read-only and reports grounding facts. Does NOT modify it."""
import hashlib, os, sys, re, json
import openpyxl

WB_PATH = "TTW_NFL_v1_1_1 Version 2.xlsx"

def sha256(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()

def main():
    facts = {}
    facts["file"] = WB_PATH
    facts["exists"] = os.path.exists(WB_PATH)
    facts["size_bytes"] = os.path.getsize(WB_PATH)
    facts["sha256"] = sha256(WB_PATH)

    # Load with formulas (data_only=False)
    wb = openpyxl.load_workbook(WB_PATH, data_only=False, keep_vba=False)
    sheets = wb.sheetnames
    facts["sheet_count"] = len(sheets)
    facts["sheet_order"] = sheets

    states = {}
    for ws in wb.worksheets:
        states[ws.title] = ws.sheet_state  # visible / hidden / veryHidden
    facts["sheet_states"] = states
    facts["hidden_sheets"] = [t for t, s in states.items() if s != "visible"]
    facts["visible_sheets"] = [t for t, s in states.items() if s == "visible"]

    # Count formula cells per sheet
    formula_counts = {}
    total_formulas = 0
    for ws in wb.worksheets:
        cnt = 0
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and v.startswith("="):
                    cnt += 1
                elif c.data_type == "f":
                    cnt += 1
        formula_counts[ws.title] = cnt
        total_formulas += cnt
    facts["formula_counts_by_sheet"] = formula_counts
    facts["total_formula_cells"] = total_formulas

    # Dimensions per sheet
    dims = {}
    for ws in wb.worksheets:
        dims[ws.title] = {"max_row": ws.max_row, "max_col": ws.max_column, "dimensions": ws.dimensions}
    facts["sheet_dimensions"] = dims

    print(json.dumps(facts, indent=2))

if __name__ == "__main__":
    main()
