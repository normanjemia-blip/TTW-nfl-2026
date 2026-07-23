#!/usr/bin/env python3
import openpyxl
WB="TTW_NFL_v1_1_1 Version 2.xlsx"
wbf=openpyxl.load_workbook(WB, data_only=False)
wbv=openpyxl.load_workbook(WB, data_only=True)
wf=wbf["MARKET LINES"]; wv=wbv["MARKET LINES"]
print("Formula in R5 (Input check):", wf.cell(5,18).value)
print("Formula in Q5 (Mkt home spread):", wf.cell(5,17).value)
print("Formula in S5 (Stale?):", wf.cell(5,19).value)
print("\nComputed Q (mkt home spread), R (input check), S (stale) for sample rows 261-276:")
for r in range(261,277):
    print(f"R{r}: Fav={wv.cell(r,7).value!r} Home={wv.cell(r,6).value!r} Q={wv.cell(r,17).value!r} R={wv.cell(r,18).value!r} S={wv.cell(r,19).value!r}")

# ENGINE: how does it look up market lines? find formulas referencing 'MARKET LINES' or named ranges
print("\n--- ENGINE header row & sample formula referencing market ---")
we=wbf["ENGINE"]
for r in range(1,7):
    cells=[f"{we.cell(r,c).coordinate}={repr(we.cell(r,c).value)[:30]}" for c in range(1,20) if we.cell(r,c).value is not None]
    if cells: print(f"R{r}:"," | ".join(cells))
