#!/usr/bin/env python3
import openpyxl, datetime
WB="TTW_NFL_v1_1_1 Version 2.xlsx"
wb=openpyxl.load_workbook(WB, data_only=False)

def is_input(v):
    # A manual input = a literal (non-formula) non-empty value
    if v is None: return False
    if isinstance(v,str):
        if v.strip()=="" : return False
        if v.startswith("="): return False
        return True
    return True  # numbers, dates

# MARKET LINES manual entry cols: G,H,I,J,K (7..11), data rows 5..end
ws=wb["MARKET LINES"]
ml=[]
for r in range(5, ws.max_row+1):
    for c in (7,8,9,10,11):
        v=ws.cell(r,c).value
        if is_input(v):
            ml.append((ws.cell(r,c).coordinate, v))
print(f"MARKET LINES manual entries (Favorite/Spread/Total/Open cols): {len(ml)}")
for x in ml[:10]: print("   ", x)

# ADJUSTMENTS input cols A,B,C,D,G,H,I,J,K,L (1,2,3,4,7,8,9,10,11,12) rows 5..end
ws=wb["ADJUSTMENTS"]
adj=[]
for r in range(5, ws.max_row+1):
    for c in (1,2,3,4,7,8,9,10,11,12):
        v=ws.cell(r,c).value
        if is_input(v):
            adj.append((ws.cell(r,c).coordinate, v))
print(f"\nADJUSTMENTS manual entries: {len(adj)}")
for x in adj[:10]: print("   ", x)

# QB VALUES: check Active value (E) vs Baseline (C) mismatch => nonzero delta
ws=wb["QB VALUES"]
qb_nonzero=[]
qb_rows=0
for r in range(5, ws.max_row+1):
    team=ws.cell(r,1).value
    if not is_input(team): continue
    qb_rows+=1
    base=ws.cell(r,3).value
    act=ws.cell(r,5).value
    if isinstance(base,(int,float)) and isinstance(act,(int,float)) and base!=act:
        qb_nonzero.append((team, base, act, round(act-base,3)))
print(f"\nQB VALUES team rows: {qb_rows}; rows with nonzero delta (Active != Baseline): {len(qb_nonzero)}")
for x in qb_nonzero: print("   ", x)

# TEAM RATINGS override col I (9), rows 5..36
ws=wb["TEAM RATINGS"]
ovr=[]
for r in range(5, ws.max_row+1):
    v=ws.cell(r,9).value
    if is_input(v):
        ovr.append((ws.cell(r,9).coordinate, v))
print(f"\nTEAM RATINGS manual overrides (col I): {len(ovr)}")
for x in ovr[:10]: print("   ", x)

# SETTINGS HFA exceptions block rows 65..? (A=team)
ws=wb["SETTINGS"]
hfa=[]
for r in range(65, 66+40):
    v=ws.cell(r,1).value
    if is_input(v):
        hfa.append((r, v, ws.cell(r,2).value))
print(f"\nSETTINGS Team HFA exceptions entered: {len(hfa)}")
for x in hfa: print("   ", x)
