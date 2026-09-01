#!/usr/bin/env python3
"""Phase 5C post-promotion production verification — READ-ONLY.

Drives every check off the promotion manifest itself: reads back all 133 directly
written cells and all 226 recalculated cells against their expected "After" values,
then confirms formula coordinates, the inherited-error budget and the Week-1 slate.

Usage: verify_production_post_promotion.py <production_export.xlsx>

Writes nothing to Google. Exits non-zero on any failure.
"""
import csv
import os
import sys

import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CAND = os.path.join(ROOT, "TTW_NFL_Power_Ratings_2026_v1.4_SRCB_BLENDFIX_CANDIDATE.xlsx")
CKPT = os.path.join(ROOT, "TTW_NFL_2026_PROD_ROLLBACK_CHECKPOINT_20260901T1432Z.xlsx")
MANIFEST = os.path.join(ROOT, "audit", "TTW_NFL_2026_Promotion_Manifest_v14_20260901.csv")
SLATE = os.path.join(ROOT, "audit", "TTW_NFL_2026_Candidate_Week1_Slate_20260901.csv")
OUT = os.path.join(ROOT, "audit", "TTW_NFL_2026_Production_Readback_20260901.csv")

ERRORS = ("#REF!", "#VALUE!", "#N/A", "#DIV/0!", "#NAME?", "#NUM!", "#NULL!", "#ERROR!")
INHERITED = {("CALC", "B%d" % r) for r in range(39, 44)} | {("DATA QUALITY", "B8")}

R = []
def check(name, ok, detail=""):
    R.append((name, bool(ok), detail))


def ftext(c):
    v = c.value
    if hasattr(v, "text"):
        return v.text.lstrip("=")
    return v[1:] if isinstance(v, str) and v.startswith("=") else None


def norm(v):
    if v is None:
        return "(blank)"
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v)


def main(prod):
    pf = openpyxl.load_workbook(prod)
    pv = openpyxl.load_workbook(prod, data_only=True)
    cf = openpyxl.load_workbook(CAND)
    cv = openpyxl.load_workbook(CAND, data_only=True)
    kv = openpyxl.load_workbook(CKPT, data_only=True)

    rows = list(csv.DictReader(open(MANIFEST)))
    written = [r for r in rows if r["Action"] in ("PASTE VALUE", "PASTE TEXT", "FORMULA", "TEXT")]
    recalc = [r for r in rows if r["Action"].startswith("RECALCULATED")
              and ":" not in r["Cell / Range"] and " " not in r["Cell / Range"]]
    check("manifest_splits_133_written_226_recalculated",
          len(written) == 133 and len([r for r in rows if r["Action"].startswith("RECALCULATED")]) == 226,
          "%d written / %d recalculated" % (len(written), len(rows) - len(written)))

    # ---------- readback: the 133 cells a human entered ----------
    audit, bad_w = [], []
    for r in written:
        sh, cell, want = r["Sheet"], r["Cell / Range"], r["After"]
        if r["Action"] == "FORMULA":
            got = "=" + (ftext(pf[sh][cell]) or "")
            ok = got == want
        elif want.endswith("…"):                      # manifest truncates long text
            got = norm(pv[sh][cell].value)
            ok = got.startswith(want[:-1])
        else:
            got = norm(pv[sh][cell].value)
            ok = got == want
        if not ok:
            bad_w.append("%s!%s got %r want %r" % (sh, cell, got[:60], want[:60]))
        audit.append([sh, cell, r["Team"], "WRITTEN", r["Before"][:40], want[:40], got[:40],
                      "OK" if ok else "MISMATCH"])
    check("readback_133_written_cells", not bad_w,
          "%d mismatched: %s" % (len(bad_w), bad_w[:4]) if bad_w else "133/133 match the manifest 'After'")

    # ---------- readback: the 226 recalculated cells ----------
    bad_r = []
    for r in recalc:
        sh, cell, want = r["Sheet"], r["Cell / Range"], r["After"]
        got = norm(pv[sh][cell].value)
        ok = got == want
        if not ok:
            bad_r.append("%s!%s got %r want %r" % (sh, cell, got, want))
        audit.append([sh, cell, r["Team"], "RECALCULATED", r["Before"][:40], want[:40], got[:40],
                      "OK" if ok else "MISMATCH"])
    check("recalculated_cells_match_expected", not bad_r,
          "%d mismatched: %s" % (len(bad_r), bad_r[:4]) if bad_r
          else "%d/%d PRESEASON + TEAM RATINGS outputs match" % (len(recalc), len(recalc)))

    # ---------- GP=0 fallback ----------
    tr, ps = pv["TEAM RATINGS"], pv["PRESEASON"]
    teams = [tr.cell(r, 1).value for r in range(5, 37)]
    check("gp_zero_all_32", all(tr.cell(r, 2).value == 0 for r in range(5, 37)))
    nb = [teams[i] for i in range(32) if tr.cell(5 + i, 4).value not in (None, "")]
    check("D5_D36_blank_not_zero", not nb, "non-blank: %s" % nb if nb else "32/32 blank")
    fb = [teams[i] for i in range(32)
          if not (tr.cell(5 + i, 6).value == tr.cell(5 + i, 8).value == tr.cell(5 + i, 10).value
                  == ps.cell(5 + i, 19).value)]
    check("gp0_fallback_F_H_J_equal_prior_all_32", not fb,
          "mismatched: %s" % fb if fb else "32/32 retain 100% of the preseason prior")

    # ---------- formula count and coordinates ----------
    fp = {(s, c.coordinate): ftext(c) for s in pf.sheetnames
          for row in pf[s].iter_rows() for c in row if ftext(c) is not None}
    fc = {(s, c.coordinate): ftext(c) for s in cf.sheetnames
          for row in cf[s].iter_rows() for c in row if ftext(c) is not None}
    check("formula_count_57399", len(fp) == 57399, str(len(fp)))
    check("formula_coordinates_identical_to_candidate", set(fp) == set(fc),
          "symmetric difference: %d" % len(set(fp) ^ set(fc)))
    ftd = sorted(k for k in fc if fp.get(k) != fc[k])
    check("formula_text_identical_to_candidate", not ftd,
          "%d differing: %s" % (len(ftd), ftd[:4]) if ftd else "57,399/57,399")

    # ---------- error budget ----------
    found = []
    for s in pv.sheetnames:
        for row in pv[s].iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and any(e in v for e in ERRORS):
                    found.append((s, c.coordinate, v[:30]))
    new = [f for f in found if (f[0], f[1]) not in INHERITED]
    check("exactly_six_inherited_errors", len(found) == 6 and not new,
          "%d total, %d new" % (len(found), len(new)))
    check("inherited_errors_are_the_documented_six",
          {(f[0], f[1]) for f in found} == INHERITED,
          "CALC B39:B43 + DATA QUALITY B8")
    kfound = sum(1 for s in kv.sheetnames for row in kv[s].iter_rows() for c in row
                 if isinstance(c.value, str) and any(e in c.value for e in ERRORS))
    check("error_count_unchanged_vs_pre_promotion_checkpoint", kfound == len(found),
          "pre %d -> post %d" % (kfound, len(found)))

    # ---------- unrelated production data preserved ----------
    for sheet, rng in (("MARKET LINES", (5, 61, 1, 19)), ("QB VALUES", (5, 36, 1, 14)),
                       ("ADJUSTMENTS", (5, 104, 1, 13)), ("PRESEASON MONITOR", (1, 40, 1, 18)),
                       ("IMPORT SCHEDULE", (5, 300, 1, 12))):
        d = [(r, c) for r in range(rng[0], rng[1] + 1) for c in range(rng[2], rng[3] + 1)
             if kv[sheet].cell(r, c).value != pv[sheet].cell(r, c).value]
        check("preserved_" + sheet.replace(" ", "_").lower(), not d,
              "%d changed" % len(d) if d else "identical to the pre-promotion checkpoint")
    check("srcC_still_blank_all_32", all(ps.cell(r, 14).value is None for r in range(5, 37)))
    st = {str(pv["SETTINGS"].cell(r, 1).value).strip(): pv["SETTINGS"].cell(r, 2).value
          for r in range(1, 80) if pv["SETTINGS"].cell(r, 1).value is not None}
    check("validate_only_and_thresholds_intact",
          st.get("Win-totals mode (VALIDATE-ONLY until conversion is verified)") == "VALIDATE-ONLY"
          and st.get("ATS BET at >=") == 1.5 and st.get("ATS INVESTIGATE at >=") == 1.5
          and st.get("ATS LEAN at >=") == 1.0
          and st.get("Enable BET labels (OFF = 1.5+ ATS edges display STRONG INVESTIGATE)") == "Y")
    bw = [teams[i] for i in range(32)
          if (ps.cell(5 + i, 8).value, ps.cell(5 + i, 13).value, ps.cell(5 + i, 18).value)
          != (0.4, 0.35, 0.25)]
    check("weights_A040_B035_C025", not bw, "offending: %s" % bw if bw else "32/32")

    # ---------- 16-game slate ----------
    slate = {r["GameID"]: r for r in csv.DictReader(open(SLATE))}
    en, da = pv["ENGINE"], pv["DASHBOARD"]
    games, gm = [], []
    for r in range(5, 300):
        gid = en.cell(r, 2).value
        if not gid or en.cell(r, 3).value != 1:
            continue
        s = slate[gid]
        got = (round(en.cell(r, 19).value, 2), en.cell(r, 20).value,
               round(en.cell(r, 22).value, 2), en.cell(r, 24).value)
        exp = (float(s["FinalMargin"]), s["Model spread (fair line)"],
               float(s["SpreadEdge"]), s["Supported side"])
        if got != exp:
            gm.append((gid, got, exp))
        games.append([gid, en.cell(r, 5).value, en.cell(r, 6).value, got[0], got[1],
                      en.cell(r, 21).value, got[2], got[3], "OK" if got == exp else "MISMATCH"])
    check("slate_reconciliation_16_of_16", len(games) == 16 and not gm,
          "%d games, mismatches %s" % (len(games), gm[:3]) if gm else "16/16")
    check("dashboard_shows_16_games",
          sum(1 for r in range(1, 40) if isinstance(da.cell(r, 1).value, str)
              and " @ " in str(da.cell(r, 1).value)) == 16)

    # ---------- pins ----------
    i_d, i_n = teams.index("DAL"), teams.index("NYG")
    dn = [g for g in games if g[0] == "2026_01_DAL_NYG"][0]
    for n, ok in (("pin_DAL_-1.07_rank_22",
                   tr.cell(5 + i_d, 10).value == -1.07 and tr.cell(5 + i_d, 11).value == 22),
                  ("pin_NYG_-1.90_rank_24",
                   tr.cell(5 + i_n, 10).value == -1.9 and tr.cell(5 + i_n, 11).value == 24),
                  ("pin_final_margin_+0.77", dn[3] == 0.77),
                  ("pin_fair_line_NYG_-0.8", dn[4] == "NYG -0.8"),
                  ("pin_edge_+3.27_on_NYG_+2.5", dn[6] == 3.27 and dn[7].startswith("NYG"))):
        check(n, ok)

    with open(OUT, "w", newline="") as f:
        w = csv.writer(f, lineterminator="\n")
        w.writerow(["Sheet", "Cell", "Team", "Class", "Before", "Expected after", "Read back", "Result"])
        w.writerows(audit)
        w.writerow([])
        w.writerow(["GameID", "Away", "Home", "FinalMargin", "Fair line", "MktHomeSpr",
                    "Edge", "Side", "Result"])
        w.writerows(games)

    width = max(len(n) for n, _o, _d in R)
    for n, ok, d in R:
        print("%-4s %-*s %s" % ("PASS" if ok else "FAIL", width, n, d))
    bad_n = [n for n, ok, _d in R if not ok]
    print("\n%d checks, %d failed" % (len(R), len(bad_n)))
    print("readback written to", OUT)
    return 1 if bad_n else 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1]))
